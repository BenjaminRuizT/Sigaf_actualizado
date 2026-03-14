from fastapi import FastAPI, APIRouter, Depends, HTTPException, Header, UploadFile, File
from fastapi.responses import StreamingResponse
from dotenv import load_dotenv
from starlette.middleware.cors import CORSMiddleware
from motor.motor_asyncio import AsyncIOMotorClient
import os, logging, uuid, io, base64, asyncio, hashlib, hmac, json
from pathlib import Path
from pydantic import BaseModel
from typing import List, Optional
from datetime import datetime, timezone, timedelta
import jwt
import bcrypt
import openpyxl
from cryptography.fernet import Fernet
from cryptography.hazmat.primitives.kdf.pbkdf2 import PBKDF2HMAC
from cryptography.hazmat.primitives import hashes as crypto_hashes
from pdf_generator import generate_user_manual, generate_presentation

ROOT_DIR = Path(__file__).parent
load_dotenv(ROOT_DIR / '.env', override=False)

SECRET_KEY = os.environ.get('JWT_SECRET', 'sigaf-jwt-secret-2024')
mongo_url = os.environ['MONGO_URL']
client = AsyncIOMotorClient(mongo_url)
db = client[os.environ['DB_NAME']]

# ── Field Encryption (Fernet / AES-128-CBC + HMAC-SHA256) ─────────────────────
# Key is derived from JWT_SECRET via PBKDF2 so it is never stored in plain text.
# A fixed application-level salt is intentional here: we need deterministic key
# derivation across restarts.  The actual secrecy comes from JWT_SECRET.
_ENC_SALT = b"sigaf-field-enc-salt-v1"

def _build_fernet() -> Fernet:
    kdf = PBKDF2HMAC(
        algorithm=crypto_hashes.SHA256(),
        length=32,
        salt=_ENC_SALT,
        iterations=100_000,
    )
    raw_key = kdf.derive(SECRET_KEY.encode())
    return Fernet(base64.urlsafe_b64encode(raw_key))

_fernet = _build_fernet()

def encrypt_field(value: str) -> str:
    """Encrypt a string field and return a base64url-encoded ciphertext prefixed with 'enc:'."""
    if not value:
        return value
    return "enc:" + _fernet.encrypt(value.encode()).decode()

def decrypt_field(value: str) -> str:
    """Decrypt a field encrypted by encrypt_field(). Returns value unchanged if not encrypted."""
    if not value or not str(value).startswith("enc:"):
        return value
    try:
        return _fernet.decrypt(value[4:].encode()).decode()
    except Exception:
        return value  # Return as-is if decryption fails (e.g. data from before encryption was enabled)

# Sensitive fields that are encrypted at rest in the equipment collection
_SENSITIVE_EQ_FIELDS = ("serie", "factura")

def encrypt_equipment(doc: dict) -> dict:
    """Return a copy of doc with sensitive equipment fields encrypted."""
    out = dict(doc)
    for f in _SENSITIVE_EQ_FIELDS:
        if f in out and out[f] and not str(out[f]).startswith("enc:"):
            out[f] = encrypt_field(str(out[f]))
    return out

def decrypt_equipment(doc: dict) -> dict:
    """Return a copy of doc with sensitive equipment fields decrypted."""
    if not doc:
        return doc
    out = dict(doc)
    for f in _SENSITIVE_EQ_FIELDS:
        if f in out:
            out[f] = decrypt_field(str(out[f]))
    return out

# ── Audit Digital Signature (HMAC-SHA256) ─────────────────────────────────────
_HMAC_KEY = hashlib.sha256(f"sigaf-audit-sign:{SECRET_KEY}".encode()).digest()

def _audit_payload(audit: dict, scans: list) -> str:
    """Build a canonical, deterministic string from audit data to sign."""
    sorted_scans = sorted(
        [{"b": s.get("codigo_barras", ""), "c": s.get("classification", "")} for s in scans],
        key=lambda x: x["b"]
    )
    payload = {
        "audit_id":        audit.get("id", ""),
        "cr_tienda":       audit.get("cr_tienda", ""),
        "auditor_id":      audit.get("auditor_id", ""),
        "started_at":      audit.get("started_at", ""),
        "finished_at":     audit.get("finished_at", ""),
        "total_equipment": audit.get("total_equipment", 0),
        "located_count":   audit.get("located_count", 0),
        "not_found_count": audit.get("not_found_count", 0),
        "not_found_value": audit.get("not_found_value", 0),
        "scans":           sorted_scans,
    }
    return json.dumps(payload, ensure_ascii=False, sort_keys=True)

def sign_audit(audit: dict, scans: list) -> dict:
    """Return a signature object {hash, algorithm, signed_at}."""
    payload_str = _audit_payload(audit, scans)
    digest = hmac.new(_HMAC_KEY, payload_str.encode(), hashlib.sha256).hexdigest()
    return {
        "hash":       digest,
        "algorithm":  "HMAC-SHA256",
        "signed_at":  datetime.now(timezone.utc).isoformat(),
        "signed_by":  audit.get("auditor_name", "system"),
    }

def verify_audit_signature(audit: dict, scans: list) -> bool:
    """Re-compute the signature and compare with the stored hash."""
    stored = (audit.get("signature") or {}).get("hash")
    if not stored:
        return False
    payload_str = _audit_payload(audit, scans)
    expected = hmac.new(_HMAC_KEY, payload_str.encode(), hashlib.sha256).hexdigest()
    return hmac.compare_digest(stored, expected)

app = FastAPI(title="SIGAF API")
api_router = APIRouter(prefix="/api")

logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(name)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)

# ==================== ADMIN HISTORY (SNAPSHOTS + ROLLBACK) ====================
# Every destructive or mutating Super Admin action saves a snapshot to db.admin_history.
# Snapshots store the BEFORE state so the action can be undone.

class HistAction:
    DELETE_AUDIT     = "DELETE_AUDIT"
    DELETE_USER      = "DELETE_USER"
    UPDATE_USER      = "UPDATE_USER"
    CREATE_USER      = "CREATE_USER"
    UPDATE_EQUIPMENT = "UPDATE_EQUIPMENT"
    DATA_RESET       = "DATA_RESET"

_ROLLBACK_SUPPORTED = {HistAction.DELETE_AUDIT, HistAction.DELETE_USER,
                       HistAction.UPDATE_USER, HistAction.UPDATE_EQUIPMENT}

async def save_history(
    action: str,
    actor_email: str,
    actor_id: str,
    target_id: Optional[str],
    target_label: Optional[str],
    before: Optional[dict],
    extra: Optional[dict] = None,
) -> str:
    """Persist a history snapshot; returns the generated snapshot_id."""
    snapshot_id = str(uuid.uuid4())
    try:
        doc = {
            "id": snapshot_id,
            "ts": datetime.now(timezone.utc).isoformat(),
            "action": action,
            "actor_email": actor_email,
            "actor_id": actor_id,
            "target_id": target_id,
            "target_label": target_label,
            "before": before,
            "extra": extra or {},
            "rolled_back": False,
            "rolled_back_at": None,
            "rolled_back_by": None,
            "can_rollback": action in _ROLLBACK_SUPPORTED,
        }
        await db.admin_history.insert_one(doc)
    except Exception:
        pass  # history must never crash business logic
    return snapshot_id

# ==================== SECURITY AUDIT LOGS ====================
# A structured event log for security-relevant actions, distinct from
# the HTTP request log (app_logs).  Never raises — logging must never
# block or crash a business operation.

class SecEvent:
    """Constants for security event types."""
    LOGIN_SUCCESS    = "LOGIN_SUCCESS"
    LOGIN_FAILED     = "LOGIN_FAILED"
    ACCOUNT_LOCKED   = "ACCOUNT_LOCKED"
    ACCOUNT_UNLOCKED = "ACCOUNT_UNLOCKED"
    UNLOCK_REQUESTED = "UNLOCK_REQUESTED"
    PASSWORD_CHANGED = "PASSWORD_CHANGED"
    USER_CREATED     = "USER_CREATED"
    USER_UPDATED     = "USER_UPDATED"
    USER_DELETED     = "USER_DELETED"
    ROLE_CHANGED     = "ROLE_CHANGED"
    AUDIT_FINALIZED  = "AUDIT_FINALIZED"
    AUDIT_DELETED    = "AUDIT_DELETED"
    AUDIT_CANCELLED  = "AUDIT_CANCELLED"
    AUDIT_RESTORED   = "AUDIT_RESTORED"
    SESSION_CLOSED   = "SESSION_CLOSED"
    EQUIPMENT_EDITED = "EQUIPMENT_EDITED"
    DATA_RESET       = "DATA_RESET"
    SIGNATURE_VALID   = "SIGNATURE_VALID"
    SIGNATURE_INVALID = "SIGNATURE_INVALID"

class SecLevel:
    INFO     = "INFO"
    WARNING  = "WARNING"
    CRITICAL = "CRITICAL"

_SEC_LEVEL_MAP = {
    SecEvent.LOGIN_SUCCESS:    SecLevel.INFO,
    SecEvent.LOGIN_FAILED:     SecLevel.WARNING,
    SecEvent.ACCOUNT_LOCKED:   SecLevel.CRITICAL,
    SecEvent.ACCOUNT_UNLOCKED: SecLevel.WARNING,
    SecEvent.UNLOCK_REQUESTED: SecLevel.WARNING,
    SecEvent.PASSWORD_CHANGED: SecLevel.WARNING,
    SecEvent.USER_CREATED:     SecLevel.WARNING,
    SecEvent.USER_UPDATED:     SecLevel.WARNING,
    SecEvent.USER_DELETED:     SecLevel.CRITICAL,
    SecEvent.ROLE_CHANGED:     SecLevel.CRITICAL,
    SecEvent.AUDIT_FINALIZED:  SecLevel.INFO,
    SecEvent.AUDIT_DELETED:    SecLevel.CRITICAL,
    SecEvent.AUDIT_CANCELLED:  SecLevel.WARNING,
    SecEvent.AUDIT_RESTORED:   SecLevel.WARNING,
    SecEvent.SESSION_CLOSED:   SecLevel.WARNING,
    SecEvent.EQUIPMENT_EDITED: SecLevel.WARNING,
    SecEvent.DATA_RESET:       SecLevel.CRITICAL,
    SecEvent.SIGNATURE_VALID:  SecLevel.INFO,
    SecEvent.SIGNATURE_INVALID: SecLevel.CRITICAL,
}

async def sec_log(
    event: str,
    actor_email: Optional[str] = None,
    actor_id: Optional[str] = None,
    target: Optional[str] = None,
    detail: Optional[dict] = None,
    ip: Optional[str] = None,
) -> None:
    """Insert one security event into db.security_logs.  Never raises."""
    try:
        await db.security_logs.insert_one({
            "ts":          datetime.now(timezone.utc).isoformat(),
            "event":       event,
            "level":       _SEC_LEVEL_MAP.get(event, SecLevel.INFO),
            "actor_email": actor_email,
            "actor_id":    actor_id,
            "target":      target,
            "detail":      detail or {},
            "ip":          ip,
        })
    except Exception:
        pass  # Security logging must never crash business logic

# ==================== REQUEST LOGGING MIDDLEWARE ====================
import time as _time
from starlette.middleware.base import BaseHTTPMiddleware
from starlette.requests import Request as StarletteRequest

class AppLogMiddleware(BaseHTTPMiddleware):
    """Log every API request to MongoDB for Super Admin visibility."""
    SKIP_PATHS = {"/api/admin/app-logs", "/api/auth/me"}  # avoid infinite loops & noise

    async def dispatch(self, request: StarletteRequest, call_next):
        if request.url.path in self.SKIP_PATHS:
            return await call_next(request)

        start = _time.monotonic()
        error_detail = None
        status_code = 200
        try:
            response = await call_next(request)
            status_code = response.status_code
            return response
        except Exception as exc:
            status_code = 500
            error_detail = str(exc)
            raise
        finally:
            duration_ms = round((_time.monotonic() - start) * 1000)
            is_error = status_code >= 400
            # Extract user from Authorization header if present
            auth_header = request.headers.get("Authorization", "")
            user_email = None
            if auth_header.startswith("Bearer "):
                try:
                    import jwt as _jwt
                    SECRET_KEY_LOCAL = os.environ.get("JWT_SECRET", "sigaf-secret-key-2024")
                    payload = _jwt.decode(auth_header.split(" ")[1], SECRET_KEY_LOCAL, algorithms=["HS256"])
                    user_email = payload.get("email")
                except Exception:
                    pass

            log_doc = {
                "ts": datetime.now(timezone.utc).isoformat(),
                "method": request.method,
                "path": request.url.path,
                "status": status_code,
                "duration_ms": duration_ms,
                "user_email": user_email,
                "ip": request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown"),
                "is_error": is_error,
                "error_detail": error_detail,
                "user_agent": request.headers.get("User-Agent", "")[:200],
            }
            try:
                await db.app_logs.insert_one(log_doc)
            except Exception:
                pass  # Never let logging break the app

# ==================== MODELS ====================
class LoginInput(BaseModel):
    email: str
    password: str

class ScanInput(BaseModel):
    barcode: str

class AuditCreateInput(BaseModel):
    cr_tienda: str

class MovementInput(BaseModel):
    audit_id: str
    equipment_id: str
    type: str
    from_cr_tienda: Optional[str] = None
    to_cr_tienda: Optional[str] = None

class UserCreateInput(BaseModel):
    nombre: str
    email: str
    password: str
    perfil: str

class UserUpdateInput(BaseModel):
    nombre: Optional[str] = None
    email: Optional[str] = None
    password: Optional[str] = None
    perfil: Optional[str] = None

class EquipmentUpdateInput(BaseModel):
    codigo_barras: Optional[str] = None
    no_activo: Optional[str] = None
    descripcion: Optional[str] = None
    marca: Optional[str] = None
    modelo: Optional[str] = None
    serie: Optional[str] = None
    costo: Optional[float] = None
    depreciacion: Optional[float] = None

class AuditCancelInput(BaseModel):
    reason: str

class UnknownSurplusInput(BaseModel):
    codigo_barras: str
    descripcion: str
    marca: str
    modelo: str
    no_activo: Optional[str] = ""
    serie: Optional[str] = ""

class FinalizeWithPhotosInput(BaseModel):
    photo_ab_base64: Optional[str] = None    # Photo for ALTA/BAJA movements
    photo_transfer_base64: Optional[str] = None  # Photo for TRANSFER movements

# ==================== AUTH HELPERS ====================
def hash_password(password: str) -> str:
    return bcrypt.hashpw(password.encode('utf-8'), bcrypt.gensalt()).decode('utf-8')

def verify_password(password: str, hashed: str) -> bool:
    return bcrypt.checkpw(password.encode('utf-8'), hashed.encode('utf-8'))

def create_token(user_id: str, email: str, perfil: str, nombre: str, session_id: str = "") -> str:
    from datetime import timedelta
    expire = datetime.now(timezone.utc) + timedelta(hours=24)
    return jwt.encode({"sub": user_id, "email": email, "perfil": perfil, "nombre": nombre,
                       "sid": session_id, "exp": expire}, SECRET_KEY, algorithm="HS256")

# ==================== ACCOUNT LOCKOUT CONSTANTS ====================
_LOGIN_MAX_ATTEMPTS = 5     # consecutive failures before lock
_LOGIN_LOCK_SECONDS = 1800  # 30 minutes auto-unlock window

def _seconds_since_iso(iso_str: str) -> float:
    """Safe datetime diff helper — handles both tz-aware and tz-naive ISO strings."""
    try:
        dt = datetime.fromisoformat(iso_str)
        now = datetime.now(timezone.utc)
        # Normalize: if dt has no tzinfo, assume UTC
        if dt.tzinfo is None:
            dt = dt.replace(tzinfo=timezone.utc)
        return (now - dt).total_seconds()
    except Exception:
        return 0.0

async def get_current_user(authorization: Optional[str] = Header(None)):
    """Decode JWT, verify account is not locked, and verify session is still active."""
    if not authorization or not authorization.startswith("Bearer "):
        raise HTTPException(401, "Not authenticated")
    try:
        payload = jwt.decode(authorization.split(" ")[1], SECRET_KEY, algorithms=["HS256"])
    except jwt.ExpiredSignatureError:
        raise HTTPException(401, "Token expired")
    except jwt.InvalidTokenError:
        raise HTTPException(401, "Invalid token")

    # Verify account lock status on every authenticated request
    user_doc = await db.users.find_one({"id": payload["sub"]}, {"_id": 0, "locked": 1, "locked_at": 1})
    if user_doc and user_doc.get("locked"):
        locked_at_str = user_doc.get("locked_at") or ""
        elapsed = _seconds_since_iso(locked_at_str) if locked_at_str else _LOGIN_LOCK_SECONDS + 1
        if elapsed >= _LOGIN_LOCK_SECONDS:
            await db.users.update_one(
                {"id": payload["sub"]},
                {"$set": {"locked": False, "failed_login_attempts": 0, "locked_at": None,
                           "unlock_requested": False}}
            )
        else:
            raise HTTPException(401, "Account locked")

    # Validate session is still active (session_id embedded in JWT)
    sid = payload.get("sid")
    if sid:
        session = await db.active_sessions.find_one({"id": sid, "user_id": payload["sub"]})
        if not session:
            raise HTTPException(401, "Session closed")
        # Update last_seen (fire-and-forget — don't block request)
        await db.active_sessions.update_one(
            {"id": sid}, {"$set": {"last_seen": datetime.now(timezone.utc).isoformat()}}
        )

    return payload

# ==================== ACTIVE SESSIONS ====================
# Each login creates a session doc in db.active_sessions.
# get_current_user validates the session_id embedded in the JWT.
# Only one session per user is allowed at a time.

async def create_session(user_id: str, email: str, perfil: str,
                         ip: str = "unknown", user_agent: str = "") -> str:
    """Create a new session record and return its session_id."""
    session_id = str(uuid.uuid4())
    now = datetime.now(timezone.utc).isoformat()
    await db.active_sessions.insert_one({
        "id": session_id,
        "user_id": user_id,
        "email": email,
        "perfil": perfil,
        "created_at": now,
        "last_seen": now,
        "ip": ip,
        "user_agent": user_agent[:200] if user_agent else "",
    })
    return session_id

async def close_sessions(user_id: str, exclude_session_id: Optional[str] = None):
    """Delete all active sessions for a user, optionally keeping one."""
    query: dict = {"user_id": user_id}
    if exclude_session_id:
        query["id"] = {"$ne": exclude_session_id}
    await db.active_sessions.delete_many(query)

async def count_active_sessions(user_id: str) -> int:
    return await db.active_sessions.count_documents({"user_id": user_id})

# ==================== AUTH ROUTES ====================

class UnlockRequestInput(BaseModel):
    email: str
    reason: Optional[str] = None

async def _record_failed_login(user_id: str, now_iso: str) -> int:
    """Increment failed attempts. Returns new attempt count."""
    user_doc = await db.users.find_one({"id": user_id}, {"_id": 0, "failed_login_attempts": 1})
    new_attempts = (user_doc.get("failed_login_attempts") or 0) + 1 if user_doc else 1
    update: dict = {"failed_login_attempts": new_attempts, "last_failed_at": now_iso}
    if new_attempts >= _LOGIN_MAX_ATTEMPTS:
        update.update({"locked": True, "locked_at": now_iso, "unlock_requested": False,
                        "unlock_request_reason": None, "unlock_request_at": None})
    await db.users.update_one({"id": user_id}, {"$set": update})
    return new_attempts

@api_router.post("/auth/login")
async def login(input: LoginInput, request: StarletteRequest):
    email_key = input.email.strip().lower()
    # Try exact match first, then case-insensitive fallback
    user = await db.users.find_one({"email": email_key}, {"_id": 0})
    if not user:
        user = await db.users.find_one({"email": input.email.strip()}, {"_id": 0})
    if not user:
        raise HTTPException(401, "Credenciales inválidas")

    now_iso = datetime.now(timezone.utc).isoformat()

    # Check and auto-unlock if lock window expired
    if user.get("locked"):
        locked_at_str = user.get("locked_at") or ""
        elapsed = _seconds_since_iso(locked_at_str) if locked_at_str else _LOGIN_LOCK_SECONDS + 1
        if elapsed >= _LOGIN_LOCK_SECONDS:
            await db.users.update_one(
                {"id": user["id"]},
                {"$set": {"locked": False, "failed_login_attempts": 0, "locked_at": None,
                           "unlock_requested": False}}
            )
            user = {**user, "locked": False, "failed_login_attempts": 0}
        else:
            remaining_min = max(1, int((_LOGIN_LOCK_SECONDS - elapsed) / 60))
            # Use 403 Forbidden (not 423 — more widely supported by all proxies/servers)
            raise HTTPException(403, {
                "code": "ACCOUNT_LOCKED",
                "remaining_minutes": remaining_min,
                "unlock_requested": bool(user.get("unlock_requested")),
                "user_id": user["id"],
                "email": user["email"]
            })

    if not verify_password(input.password, user["password_hash"]):
        new_attempts = await _record_failed_login(user["id"], now_iso)
        if new_attempts >= _LOGIN_MAX_ATTEMPTS:
            await sec_log(SecEvent.ACCOUNT_LOCKED, actor_email=user["email"], actor_id=user["id"],
                          detail={"reason": "max_attempts_exceeded", "attempts": new_attempts})
            raise HTTPException(403, {
                "code": "ACCOUNT_LOCKED",
                "remaining_minutes": _LOGIN_LOCK_SECONDS // 60,
                "unlock_requested": False,
                "user_id": user["id"],
                "email": user["email"]
            })
        await sec_log(SecEvent.LOGIN_FAILED, actor_email=user["email"], actor_id=user["id"],
                      detail={"attempt": new_attempts, "remaining": _LOGIN_MAX_ATTEMPTS - new_attempts})
        attempts_left = _LOGIN_MAX_ATTEMPTS - new_attempts
        detail = "Credenciales inválidas"
        if attempts_left <= 2:
            detail = f"Credenciales inválidas. {attempts_left} intento{'s' if attempts_left != 1 else ''} restante{'s' if attempts_left != 1 else ''} antes del bloqueo."
        raise HTTPException(401, detail)

    # Successful login — reset failed attempts counter
    await db.users.update_one(
        {"id": user["id"]},
        {"$set": {"failed_login_attempts": 0, "locked": False, "locked_at": None,
                   "unlock_requested": False, "last_login": now_iso}}
    )

    # Check for existing active sessions
    existing_sessions = await db.active_sessions.find(
        {"user_id": user["id"]}, {"_id": 0}
    ).to_list(100)

    # Capture IP and user-agent for new session metadata
    _req_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown")
    _req_ip = _req_ip.split(",")[0].strip()  # take first IP if comma-separated proxy chain
    _req_ua = request.headers.get("User-Agent", "")

    # Check multi-session setting
    settings_doc = await db.system_settings.find_one({"_id": "global"}, {"_id": 0}) or {}
    allow_multi = settings_doc.get("allow_multi_session", False)

    if existing_sessions and not allow_multi:
        # Return 409 Conflict so the frontend can show the session conflict dialog
        safe_sessions = [
            {
                "id": s["id"],
                "created_at": s.get("created_at"),
                "last_seen": s.get("last_seen"),
                "perfil": s.get("perfil"),
                "ip": s.get("ip", "desconocida"),
                "user_agent": s.get("user_agent", ""),
            }
            for s in existing_sessions
        ]
        raise HTTPException(409, {
            "code": "SESSION_CONFLICT",
            "active_sessions": safe_sessions,
            "user_id": user["id"],
            "email": user["email"],
            "perfil": user["perfil"],
            "nombre": user["nombre"],
        })

    # No conflict — create session and return token
    session_id = await create_session(user["id"], user["email"], user["perfil"], _req_ip, _req_ua)
    token = create_token(user["id"], user["email"], user["perfil"], user["nombre"], session_id)
    safe_user = {k: v for k, v in user.items() if k not in ("password_hash",)}
    await sec_log(SecEvent.LOGIN_SUCCESS, actor_email=user["email"], actor_id=user["id"],
                  detail={"perfil": user["perfil"]})
    return {"token": token, "user": safe_user}

@api_router.post("/auth/login/force")
async def login_force(input: LoginInput, request: StarletteRequest):
    """Close all existing sessions for the user and log in fresh."""
    email_key = input.email.strip().lower()
    user = await db.users.find_one({"email": email_key}, {"_id": 0})
    if not user:
        user = await db.users.find_one({"email": input.email.strip()}, {"_id": 0})
    if not user or not verify_password(input.password, user["password_hash"]):
        raise HTTPException(401, "Credenciales inválidas")
    if user.get("locked"):
        raise HTTPException(403, {"code": "ACCOUNT_LOCKED", "remaining_minutes": 30,
                                   "unlock_requested": bool(user.get("unlock_requested")),
                                   "user_id": user["id"], "email": user["email"]})
    _req_ip = request.headers.get("X-Forwarded-For", request.client.host if request.client else "unknown")
    _req_ip = _req_ip.split(",")[0].strip()
    _req_ua = request.headers.get("User-Agent", "")
    # Close ALL existing sessions, then create new one
    await close_sessions(user["id"])
    session_id = await create_session(user["id"], user["email"], user["perfil"], _req_ip, _req_ua)
    token = create_token(user["id"], user["email"], user["perfil"], user["nombre"], session_id)
    safe_user = {k: v for k, v in user.items() if k not in ("password_hash",)}
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.users.update_one({"id": user["id"]}, {"$set": {"last_login": now_iso}})
    await sec_log(SecEvent.LOGIN_SUCCESS, actor_email=user["email"], actor_id=user["id"],
                  detail={"perfil": user["perfil"], "forced": True})
    return {"token": token, "user": safe_user}

@api_router.post("/auth/logout")
async def logout(user=Depends(get_current_user)):
    """Close the current session."""
    sid = user.get("sid")
    if sid:
        await db.active_sessions.delete_one({"id": sid})
    await sec_log(SecEvent.LOGIN_SUCCESS, actor_email=user["email"], actor_id=user["sub"],
                  detail={"action": "logout"})
    return {"message": "Sesión cerrada"}

@api_router.get("/auth/sessions")
async def get_my_sessions(user=Depends(get_current_user)):
    """Return all active sessions for the current user."""
    sessions = await db.active_sessions.find(
        {"user_id": user["sub"]}, {"_id": 0}
    ).sort("created_at", -1).to_list(100)
    current_sid = user.get("sid")
    return [{"id": s["id"], "created_at": s.get("created_at"),
             "last_seen": s.get("last_seen"), "is_current": s["id"] == current_sid,
             "ip": s.get("ip", "desconocida"), "user_agent": s.get("user_agent", "")}
            for s in sessions]

@api_router.post("/auth/sessions/close-others")
async def close_other_sessions(user=Depends(get_current_user)):
    """Keep the current session, close all others."""
    sid = user.get("sid")
    await close_sessions(user["sub"], exclude_session_id=sid)
    return {"message": "Otras sesiones cerradas"}

async def request_unlock(input: UnlockRequestInput):
    """No auth required — user is locked out and requesting manual unlock."""
    email_key = input.email.strip().lower()
    user = await db.users.find_one({"email": email_key}, {"_id": 0})
    if not user:
        user = await db.users.find_one({"email": input.email.strip()}, {"_id": 0})
    # Silent success to prevent user enumeration even if not found/not locked
    if user and user.get("locked"):
        await db.users.update_one(
            {"id": user["id"]},
            {"$set": {"unlock_requested": True,
                       "unlock_request_at": datetime.now(timezone.utc).isoformat(),
                       "unlock_request_reason": (input.reason or "").strip()[:500]}}
        )
        await sec_log(SecEvent.UNLOCK_REQUESTED, actor_email=user["email"], actor_id=user["id"],
                      detail={"has_reason": bool(input.reason)})
    return {"ok": True}

@api_router.get("/admin/unlock-requests")
async def get_unlock_requests(current_user=Depends(get_current_user)):
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    locked = await db.users.find(
        {"locked": True}, {"_id": 0, "password_hash": 0}
    ).to_list(200)
    return locked

@api_router.post("/admin/unlock/{user_id}")
async def unlock_user(user_id: str, admin=Depends(get_current_user)):
    if admin["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if not target:
        raise HTTPException(404, "Usuario no encontrado")
    await db.users.update_one(
        {"id": user_id},
        {"$set": {"locked": False, "failed_login_attempts": 0, "locked_at": None,
                   "unlock_requested": False, "unlock_request_reason": None,
                   "unlock_request_at": None, "unlocked_by": admin["nombre"],
                   "unlocked_at": datetime.now(timezone.utc).isoformat()}}
    )
    await sec_log(SecEvent.ACCOUNT_UNLOCKED,
                  actor_email=admin["email"], actor_id=admin["sub"],
                  target=target["email"],
                  detail={"unlocked_by": admin["nombre"]})
    return {"message": f"Usuario {target['email']} desbloqueado"}

@api_router.get("/admin/app-logs")
async def get_app_logs(
    current_user=Depends(get_current_user),
    limit: int = 200,
    errors_only: bool = False,
    path: Optional[str] = None,
    user_email: Optional[str] = None,
):
    """Super Admin: retrieve recent application logs from the DB."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    query: dict = {}
    if errors_only:
        query["is_error"] = True
    if path:
        query["path"] = {"$regex": path, "$options": "i"}
    if user_email:
        query["user_email"] = {"$regex": user_email, "$options": "i"}
    logs = await db.app_logs.find(query, {"_id": 0}).sort("ts", -1).limit(min(limit, 1000)).to_list(1000)
    return logs

@api_router.delete("/admin/app-logs")
async def clear_app_logs(current_user=Depends(get_current_user)):
    """Super Admin: purge all application logs."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    result = await db.app_logs.delete_many({})
    return {"deleted": result.deleted_count}

@api_router.get("/admin/security-logs")
async def get_security_logs(
    current_user=Depends(get_current_user),
    limit: int = 300,
    level: Optional[str] = None,
    event: Optional[str] = None,
    actor_email: Optional[str] = None,
):
    """Super Admin: retrieve structured security event logs."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    query: dict = {}
    if level:
        query["level"] = level.upper()
    if event:
        query["event"] = {"$regex": event, "$options": "i"}
    if actor_email:
        query["actor_email"] = {"$regex": actor_email, "$options": "i"}
    logs = await db.security_logs.find(query, {"_id": 0}).sort("ts", -1).limit(min(limit, 2000)).to_list(2000)
    return logs

@api_router.delete("/admin/security-logs")
async def clear_security_logs(current_user=Depends(get_current_user)):
    """Super Admin: purge security logs older than 90 days (or all if forced)."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    from datetime import timedelta
    cutoff = (datetime.now(timezone.utc) - timedelta(days=90)).isoformat()
    result = await db.security_logs.delete_many({"ts": {"$lt": cutoff}})
    return {"deleted": result.deleted_count, "cutoff": cutoff}


# ==================== ADMIN HISTORY ENDPOINTS ====================

@api_router.get("/admin/history")
async def get_admin_history(
    action: Optional[str] = None,
    limit: int = 200,
    current_user=Depends(get_current_user)
):
    """Super Admin: list recent admin actions with snapshot data."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    query = {}
    if action:
        query["action"] = action
    items = await db.admin_history.find(query, {"_id": 0}).sort("ts", -1).limit(min(limit, 1000)).to_list(1000)
    return items

@api_router.post("/admin/history/{snapshot_id}/rollback")
async def rollback_action(snapshot_id: str, current_user=Depends(get_current_user)):
    """Super Admin: restore the before-state of a destructive action."""
    if current_user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    snap = await db.admin_history.find_one({"id": snapshot_id}, {"_id": 0})
    if not snap:
        raise HTTPException(404, "Snapshot no encontrado")
    if snap.get("rolled_back"):
        raise HTTPException(400, "Esta acción ya fue revertida")
    if not snap.get("can_rollback"):
        raise HTTPException(400, "Esta acción no soporta rollback")

    action = snap["action"]
    before = snap.get("before")
    now_iso = datetime.now(timezone.utc).isoformat()
    result_msg = ""

    if action == HistAction.UPDATE_USER and before:
        uid = snap["target_id"]
        safe = {k: v for k, v in before.items() if k not in ("_id",)}
        await db.users.update_one({"id": uid}, {"$set": safe})
        result_msg = f"Usuario {snap['target_label']} restaurado al estado anterior"

    elif action == HistAction.DELETE_USER and before:
        existing = await db.users.find_one({"id": before["id"]})
        if existing:
            raise HTTPException(409, "El usuario ya existe — no se puede restaurar")
        safe = {k: v for k, v in before.items() if k not in ("_id",)}
        await db.users.insert_one(safe)
        result_msg = f"Usuario {snap['target_label']} restaurado"

    elif action == HistAction.DELETE_AUDIT and before:
        existing = await db.audits.find_one({"id": before["id"]})
        if existing:
            raise HTTPException(409, "La auditoría ya existe — no se puede restaurar")
        audit_doc = {k: v for k, v in before.items() if k not in ("_id",)}
        await db.audits.insert_one(audit_doc)
        # Restore scans and movements from extra
        scans = snap.get("extra", {}).get("scans", [])
        movements = snap.get("extra", {}).get("movements", [])
        if scans:
            await db.audit_scans.insert_many([{k: v for k, v in s.items() if k != "_id"} for s in scans])
        if movements:
            await db.movements.insert_many([{k: v for k, v in m.items() if k != "_id"} for m in movements])
        # Restore store status
        cr = before.get("cr_tienda")
        if cr:
            await db.stores.update_one({"cr_tienda": cr}, {"$set": {
                "audited": before.get("status") == "completed",
                "last_audit_date": before.get("finished_at"),
                "last_audit_id": before.get("id"),
                "audit_status": before.get("status")
            }})
        result_msg = f"Auditoría {snap['target_label']} restaurada con {len(scans)} escaneos y {len(movements)} movimientos"

    elif action == HistAction.UPDATE_EQUIPMENT and before:
        eid = snap["target_id"]
        safe = {k: v for k, v in before.items() if k not in ("_id",)}
        await db.equipment.update_one({"id": eid}, {"$set": safe})
        result_msg = f"Equipo {snap['target_label']} restaurado al estado anterior"

    else:
        raise HTTPException(400, "Rollback no implementado para esta acción")

    await db.admin_history.update_one({"id": snapshot_id}, {"$set": {
        "rolled_back": True,
        "rolled_back_at": now_iso,
        "rolled_back_by": current_user["email"]
    }})
    return {"message": result_msg, "snapshot_id": snapshot_id}

# ==================== SYSTEM SETTINGS ====================
_DEFAULT_SETTINGS = {
    "photo_required_alta": True,          # Pedir foto de formato ALTAS
    "photo_required_baja": True,          # Pedir foto de formato BAJAS
    "photo_required_transf": True,        # Pedir foto de formato TRANSFERENCIAS
    "pending_photos_ttl_hours": 24,       # Horas para completar fotos antes de eliminar la auditoría
    "session_timeout_minutes": 15,        # Minutos de inactividad antes de cerrar sesión automáticamente
    "allow_multi_session": False,         # Permitir múltiples sesiones simultáneas por usuario
}

@api_router.get("/admin/system-settings")
async def get_system_settings(user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    doc = await db.system_settings.find_one({"_id": "global"})
    if not doc:
        return _DEFAULT_SETTINGS.copy()
    return {k: doc.get(k, v) for k, v in _DEFAULT_SETTINGS.items()}

@api_router.put("/admin/system-settings")
async def update_system_settings(settings: dict, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    # Cast each field to its correct type (booleans as bool, ttl_hours as int)
    INT_FIELDS = {"pending_photos_ttl_hours", "session_timeout_minutes"}
    # Clamp ranges per integer field
    INT_CLAMPS = {
        "pending_photos_ttl_hours": (1, 168),
        "session_timeout_minutes": (5, 480),
    }
    update = {}
    for k in _DEFAULT_SETTINGS:
        if k not in settings:
            continue
        if k in INT_FIELDS:
            try:
                v = int(settings[k])
                lo, hi = INT_CLAMPS.get(k, (1, 9999))
                update[k] = max(lo, min(hi, v))
            except (TypeError, ValueError):
                pass
        else:
            update[k] = bool(settings[k])
    if not update:
        raise HTTPException(400, "Sin campos válidos")
    await db.system_settings.update_one(
        {"_id": "global"}, {"$set": update}, upsert=True
    )
    await save_history("UPDATE_SETTINGS", user["email"], user["sub"],
                       None, "Configuración del sistema", None, update)
    # Return current full settings from DB
    doc = await db.system_settings.find_one({"_id": "global"}) or {}
    return {k: doc.get(k, v) for k, v in _DEFAULT_SETTINGS.items()}

@api_router.get("/system-settings/public")
async def get_public_settings(user=Depends(get_current_user)):
    """All profiles: get settings needed for audit flow."""
    doc = await db.system_settings.find_one({"_id": "global"})
    if not doc:
        return _DEFAULT_SETTINGS.copy()
    return {k: doc.get(k, v) for k, v in _DEFAULT_SETTINGS.items()}


@api_router.get("/catalog/equipment-types")
async def get_equipment_catalog(user=Depends(get_current_user)):
    """Get dynamic catalog of descriptions and brands from the MAF + curated additions."""
    # Base curated catalog covering all known OXXO equipment types
    CURATED_DESCRIPTIONS = [
        "ACCESS POINT", "CAMARA CCTV", "COMPUTADORA", "COMPUTADORA ALL IN ONE",
        "DISCO DURO EXTERNO", "DVR/NVR", "ESCANER ID / LECTOR BIOMETRICO",
        "IMPRESORA", "IMPRESORA FISCAL", "LAPTOP",
        "LECTOR DE CODIGO DE BARRAS", "LECTOR DE HUELLA", "MONITOR",
        "MONITOR P/PUNTO DE VENTA", "NOBREAK / UPS", "PROYECTOR",
        "PUNTO DE VENTA", "REGULADOR DE ENERGIA", "ROUTER",
        "SCANNER / ESCANER", "SERVIDOR", "SWITCH", "TABLET",
        "TECLADO / MOUSE", "TELEFONO IP", "TERMINAL BANCARIA",
        "REGULADOR/UPS/BATERIA", "HANDHELD P/PROCESOS", "OTRO"
    ]
    CURATED_BRANDS = [
        "APC", "APPLE", "ARUBA", "AXIS", "BELKIN", "BIXOLON",
        "BOSCH", "BROTHER", "CANON", "CISCO", "CPS", "D-LINK",
        "DAHUA", "DATALOGIC", "DELL", "EATON", "EPSON", "FANVIL",
        "FUJITSU", "GRANDSTREAM", "HANWHA", "HIKVISION", "HONEYWELL",
        "HP", "HUAWEI", "IBM", "INGENICO", "KOBLENZ", "LENOVO",
        "LEXMARK", "LG", "LOGITECH", "MERAKI", "MICROSOFT", "MOTOROLA",
        "NEWLAND", "OPTOMA", "PAX", "POLY", "REASA", "REOLINK",
        "SAMSUNG", "SAM4S", "SEAGATE", "STAR", "SYMETRY",
        "TOSHIBA", "TP-LINK", "TRIPP LITE", "UBIQUITI", "VERIFONE",
        "VERKADA", "WD", "YEALINK", "ZEBRA", "OTRO"
    ]
    # Also pull distinct values from the MAF for anything not in curated list
    pipeline_desc = [{"$group": {"_id": "$descripcion"}}, {"$match": {"_id": {"$ne": None, "$ne": ""}}}, {"$limit": 500}]
    pipeline_marca = [{"$group": {"_id": "$marca"}}, {"$match": {"_id": {"$ne": None, "$ne": ""}}}, {"$limit": 500}]
    maf_descs, maf_marcas = await asyncio.gather(
        db.equipment.aggregate(pipeline_desc).to_list(500),
        db.equipment.aggregate(pipeline_marca).to_list(500),
    )
    all_desc = sorted(set(CURATED_DESCRIPTIONS) | {d["_id"].upper().strip() for d in maf_descs if d.get("_id")})
    all_marca = sorted(set(CURATED_BRANDS) | {m["_id"].upper().strip() for m in maf_marcas if m.get("_id")})
    # Move OTRO to end
    if "OTRO" in all_desc: all_desc.remove("OTRO"); all_desc.append("OTRO")
    if "OTRO" in all_marca: all_marca.remove("OTRO"); all_marca.append("OTRO")
    return {"descriptions": all_desc, "brands": all_marca}


@api_router.get("/audits/{audit_id}/cross-analysis")
async def cross_analysis(audit_id: str, user=Depends(get_current_user)):
    """Super Admin: compare no_localizado vs sobrante for potential matches (description, brand, model, value)."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Solo Super Administrador")
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    scans = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000)
    no_loc = [s for s in scans if s.get("classification") == "no_localizado"]
    sobrante = [s for s in scans if s.get("classification") in ("sobrante", "sobrante_desconocido")]
    suggestions = _build_cross_suggestions(no_loc, sobrante, audits_map={audit_id: audit})
    return {"suggestions": suggestions, "total": len(suggestions),
            "no_localizado_count": len(no_loc), "sobrante_count": len(sobrante)}


def _build_cross_suggestions(no_loc, sobrante, audits_map=None):
    """Core matching logic shared by per-audit and global cross-analysis."""
    suggestions = []
    for nl in no_loc:
        nl_eq = nl.get("equipment_data") or {}
        nl_desc  = (nl_eq.get("descripcion") or "").upper()
        nl_marca = (nl_eq.get("marca") or "").upper()
        nl_modelo = (nl_eq.get("modelo") or "").upper()
        nl_valor  = float(nl_eq.get("valor_real") or 0)
        nl_audit  = (audits_map or {}).get(nl.get("audit_id"), {})
        for sb in sobrante:
            # Skip same scan
            if nl.get("id") == sb.get("id"):
                continue
            sb_eq = sb.get("equipment_data") or {}
            sb_desc  = (sb_eq.get("descripcion") or "").upper()
            sb_marca = (sb_eq.get("marca") or "").upper()
            sb_modelo = (sb_eq.get("modelo") or "").upper()
            sb_valor  = float(sb_eq.get("valor_real") or 0)
            sb_audit  = (audits_map or {}).get(sb.get("audit_id"), {})
            score = 0
            matches = []
            if nl_desc and sb_desc and nl_desc == sb_desc:
                score += 40; matches.append("Descripción")
            elif nl_desc and sb_desc and (nl_desc in sb_desc or sb_desc in nl_desc):
                score += 20; matches.append("Descripción (parcial)")
            if nl_marca and sb_marca and nl_marca == sb_marca:
                score += 25; matches.append("Marca")
            if nl_modelo and sb_modelo and nl_modelo == sb_modelo:
                score += 25; matches.append("Modelo")
            elif nl_modelo and sb_modelo and (nl_modelo in sb_modelo or sb_modelo in nl_modelo):
                score += 15; matches.append("Modelo (parcial)")
            if nl_valor > 0 and sb_valor > 0 and abs(nl_valor - sb_valor) / max(nl_valor, sb_valor) < 0.05:
                score += 10; matches.append("Valor similar")
            if score >= 40:
                suggestions.append({
                    "score": score,
                    "confidence": "Alta" if score >= 75 else "Media" if score >= 50 else "Baja",
                    "matches": matches,
                    "no_localizado": {
                        "scan_id": nl.get("id"),
                        "audit_id": nl.get("audit_id"),
                        "tienda": nl_audit.get("tienda", ""),
                        "plaza": nl_audit.get("plaza", ""),
                        "codigo_barras": nl.get("codigo_barras"),
                        "descripcion": nl_eq.get("descripcion"),
                        "marca": nl_eq.get("marca"),
                        "modelo": nl_eq.get("modelo"),
                        "serie": nl_eq.get("serie"),
                        "valor_real": nl_valor,
                        "no_activo": nl_eq.get("no_activo"),
                    },
                    "sobrante": {
                        "scan_id": sb.get("id"),
                        "audit_id": sb.get("audit_id"),
                        "tienda": sb_audit.get("tienda", ""),
                        "plaza": sb_audit.get("plaza", ""),
                        "codigo_barras": sb.get("codigo_barras"),
                        "descripcion": sb_eq.get("descripcion") or sb.get("descripcion_manual"),
                        "marca": sb_eq.get("marca") or sb.get("marca_manual"),
                        "modelo": sb_eq.get("modelo") or sb.get("modelo_manual"),
                        "serie": sb_eq.get("serie") or sb.get("serie_manual"),
                        "valor_real": sb_valor,
                    }
                })
    # ── Deduplicate: each no_localizado and each sobrante can appear in at most ONE pair ──────
    # Sort by score descending first so the highest-confidence match wins
    suggestions.sort(key=lambda x: x["score"], reverse=True)
    used_no_loc = set()
    used_sobrante = set()
    deduped = []
    for s in suggestions:
        nl_id = s["no_localizado"]["scan_id"]
        sb_id = s["sobrante"]["scan_id"]
        # Skip if either item was already matched in a higher-score suggestion
        if nl_id in used_no_loc or sb_id in used_sobrante:
            continue
        used_no_loc.add(nl_id)
        used_sobrante.add(sb_id)
        deduped.append(s)
    return deduped


@api_router.get("/cross-analysis/global")
async def global_cross_analysis(
    plaza: Optional[str] = None,
    limit: int = 2000,
    user=Depends(get_current_user)
):
    """Super Admin: global cross-analysis across ALL audits — no_localizado vs sobrante."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Solo Super Administrador")

    # Fetch completed/incompleto audits (optionally filtered by plaza)
    audit_q = {"status": {"$in": ["completed", "incompleto", "pending_photos"]}}
    if plaza:
        audit_q["plaza"] = plaza
    audits = await db.audits.find(audit_q, {"_id": 0, "id": 1, "tienda": 1, "plaza": 1, "cr_tienda": 1, "finished_at": 1}).to_list(5000)
    if not audits:
        return {"suggestions": [], "total": 0, "no_localizado_count": 0, "sobrante_count": 0,
                "audits_analyzed": 0, "plazas": []}

    audit_ids = [a["id"] for a in audits]
    audits_map = {a["id"]: a for a in audits}

    # Fetch all relevant scans in parallel
    no_loc_task = db.audit_scans.find(
        {"audit_id": {"$in": audit_ids}, "classification": "no_localizado"},
        {"_id": 0}
    ).to_list(limit)
    sobrante_task = db.audit_scans.find(
        {"audit_id": {"$in": audit_ids}, "classification": {"$in": ["sobrante", "sobrante_desconocido"]}},
        {"_id": 0}
    ).to_list(limit)
    no_loc, sobrante = await asyncio.gather(no_loc_task, sobrante_task)

    suggestions = _build_cross_suggestions(no_loc, sobrante, audits_map=audits_map)

    # Distinct plazas for filter UI
    plazas = sorted({a.get("plaza", "") for a in audits if a.get("plaza")})

    return {
        "suggestions": suggestions,
        "total": len(suggestions),
        "no_localizado_count": len(no_loc),
        "sobrante_count": len(sobrante),
        "audits_analyzed": len(audits),
        "plazas": plazas,
    }








@api_router.post("/auth/validate-password")
async def validate_password(body: dict, user=Depends(get_current_user)):
    """Validate the current user's password."""
    pwd = body.get("password", "")
    user_doc = await db.users.find_one({"id": user["sub"]}, {"_id": 0})
    if not user_doc:
        raise HTTPException(404, "Usuario no encontrado")
    from passlib.context import CryptContext
    _ctx = CryptContext(schemes=["bcrypt"], deprecated="auto")
    if not _ctx.verify(pwd, user_doc.get("password_hash", "")):
        raise HTTPException(401, "Contraseña incorrecta")
    return {"valid": True}

@api_router.get("/auth/me")
async def get_me(user=Depends(get_current_user)):
    user_doc = await db.users.find_one({"id": user["sub"]}, {"_id": 0, "password_hash": 0})
    if not user_doc:
        raise HTTPException(404, "User not found")
    return user_doc

class ProfileUpdateInput(BaseModel):
    nombre: Optional[str] = None
    password: Optional[str] = None

@api_router.put("/auth/profile")
async def update_profile(input: ProfileUpdateInput, user=Depends(get_current_user)):
    update = {}
    if input.nombre and input.nombre.strip():
        update["nombre"] = input.nombre.strip()
    if input.password and input.password.strip():
        update["password_hash"] = hash_password(input.password.strip())
    if not update:
        raise HTTPException(400, "No fields to update")
    await db.users.update_one({"id": user["sub"]}, {"$set": update})
    updated = await db.users.find_one({"id": user["sub"]}, {"_id": 0, "password_hash": 0})
    new_token = create_token(updated["id"], updated["email"], updated["perfil"], updated["nombre"])
    if "password_hash" in update:
        await sec_log(SecEvent.PASSWORD_CHANGED, actor_email=user["email"], actor_id=user["sub"])
    return {"user": updated, "token": new_token}

# ==================== REPORTS ====================
@api_router.get("/reports/summary")
async def get_reports_summary(user=Depends(get_current_user)):
    if user["perfil"] not in ["Administrador", "Super Administrador"]:
        raise HTTPException(403, "Access denied")
    # Equipment by plaza with costs
    plaza_pipe = [{"$group": {"_id": "$plaza", "count": {"$sum": 1}, "total_cost": {"$sum": "$costo"}, "total_real": {"$sum": "$valor_real"}, "deprecated": {"$sum": {"$cond": ["$depreciado", 1, 0]}}}}]
    plaza_data = await db.equipment.aggregate(plaza_pipe).to_list(100)
    # Audit completion by plaza
    audit_pipe = [{"$match": {"status": {"$in": ["completed"]}}}, {"$group": {"_id": "$plaza", "completed": {"$sum": {"$cond": [{"$eq": ["$status", "completed"]}, 1, 0]}}, "incompleto": {"$sum": {"$cond": [{"$eq": ["$status", "incompleto"]}, 1, 0]}}, "total_not_found": {"$sum": "$not_found_count"}, "total_not_found_value": {"$sum": "$not_found_value"}}}]
    audit_data = await db.audits.aggregate(audit_pipe).to_list(100)
    # Top stores with most missing equipment
    top_missing = await db.audits.find({"status": {"$in": ["completed"]}, "not_found_count": {"$gt": 0}}, {"_id": 0, "tienda": 1, "plaza": 1, "not_found_count": 1, "not_found_value": 1, "total_equipment": 1, "finished_at": 1}).sort("not_found_count", -1).limit(20).to_list(20)
    # Equipment by year
    year_pipe = [{"$group": {"_id": "$año_adquisicion", "count": {"$sum": 1}, "cost": {"$sum": "$costo"}}}]
    year_data = await db.equipment.aggregate(year_pipe).to_list(100)
    # Movement summary
    mov_pipe = [{"$group": {"_id": "$type", "count": {"$sum": 1}, "value": {"$sum": "$equipment_data.valor_real"}}}]
    mov_data = await db.movements.aggregate(mov_pipe).to_list(10)
    return {
        "plaza_equipment": [{**p, "plaza": p["_id"]} for p in plaza_data if p["_id"]],
        "plaza_audits": [{**a, "plaza": a["_id"]} for a in audit_data if a["_id"]],
        "top_missing_stores": top_missing,
        "equipment_by_year": sorted([{"year": y["_id"], "count": y["count"], "cost": y["cost"]} for y in year_data if y["_id"]], key=lambda x: x["year"]),
        "movement_summary": {m["_id"]: {"count": m["count"], "value": round(m["value"], 2)} for m in mov_data if m["_id"]}
    }

# ==================== DASHBOARD ====================
@api_router.get("/dashboard/stats")
async def get_dashboard_stats(plaza: Optional[str] = None, user=Depends(get_current_user)):
    store_q = {}
    eq_q = {}
    audit_q = {}
    if plaza and plaza != "all":
        store_q["plaza"] = plaza
        eq_q["plaza"] = plaza
        audit_q["plaza"] = plaza

    val_pipe = [{"$match": eq_q}, {"$group": {"_id": None, "total_cost": {"$sum": "$costo"}, "total_real": {"$sum": "$valor_real"}}}]
    plaza_pipe = [{"$group": {"_id": "$plaza", "count": {"$sum": 1}}}]

    # Execute all independent queries in parallel for ~4x speed improvement
    (
        total_stores, audited_stores, total_equipment, deprecated_eq,
        val, plaza_stats,
        completed_audits, active_audits,
        most_missing, least_missing
    ) = await asyncio.gather(
        db.stores.count_documents(store_q),
        db.stores.count_documents({**store_q, "audited": True}),
        db.equipment.count_documents(eq_q),
        db.equipment.count_documents({**eq_q, "depreciado": True}),
        db.equipment.aggregate(val_pipe).to_list(1),
        db.equipment.aggregate(plaza_pipe).to_list(100),
        db.audits.count_documents({**audit_q, "status": {"$in": ["completed"]}}),
        db.audits.count_documents({**audit_q, "status": "in_progress"}),
        db.audits.find({**audit_q, "status": {"$in": ["completed"]}}, {"_id": 0}).sort("not_found_count", -1).limit(5).to_list(5),
        db.audits.find({**audit_q, "status": {"$in": ["completed"]}, "not_found_count": {"$gte": 0}}, {"_id": 0}).sort("not_found_count", 1).limit(5).to_list(5),
    )

    total_cost = val[0]["total_cost"] if val else 0
    total_real = val[0]["total_real"] if val else 0

    return {
        "total_stores": total_stores, "audited_stores": audited_stores,
        "unaudited_stores": total_stores - audited_stores,
        "total_equipment": total_equipment, "deprecated_equipment": deprecated_eq,
        "active_equipment": total_equipment - deprecated_eq,
        "total_cost": round(total_cost, 2), "total_real_value": round(total_real, 2),
        "active_audits": active_audits, "completed_audits": completed_audits,
        "equipment_by_plaza": {s["_id"]: s["count"] for s in plaza_stats if s["_id"]},
        "stores_most_missing": most_missing, "stores_least_missing": least_missing
    }

# ==================== STORES ====================
@api_router.get("/stores")
async def get_stores(plaza: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    query = {}
    if plaza and plaza != "all":
        query["plaza"] = plaza
    if search:
        query["$or"] = [
            {"cr_tienda": {"$regex": search, "$options": "i"}},
            {"tienda": {"$regex": search, "$options": "i"}}
        ]
    skip = (page - 1) * limit
    total = await db.stores.count_documents(query)
    stores = await db.stores.find(query, {"_id": 0}).sort("tienda", 1).skip(skip).limit(limit).to_list(limit)
    return {"stores": stores, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/stores/plazas")
async def get_plazas(user=Depends(get_current_user)):
    pipeline = [{"$group": {"_id": {"cr_plaza": "$cr_plaza", "plaza": "$plaza"}, "store_count": {"$sum": 1}}}, {"$sort": {"_id.plaza": 1}}]
    plazas = await db.stores.aggregate(pipeline).to_list(100)
    return [{"cr_plaza": p["_id"]["cr_plaza"], "plaza": p["_id"]["plaza"], "store_count": p["store_count"]} for p in plazas]

@api_router.get("/stores/{cr_tienda}")
async def get_store(cr_tienda: str, user=Depends(get_current_user)):
    store = await db.stores.find_one({"cr_tienda": cr_tienda}, {"_id": 0})
    if not store:
        raise HTTPException(404, "Store not found")
    return store

@api_router.get("/stores/{cr_tienda}/equipment")
async def get_store_equipment(cr_tienda: str, search: Optional[str] = None, page: int = 1, limit: int = 100, user=Depends(get_current_user)):
    query = {"cr_tienda": cr_tienda}
    if search:
        query["$or"] = [
            {"codigo_barras": {"$regex": search, "$options": "i"}},
            {"no_activo": {"$regex": search, "$options": "i"}},
            {"descripcion": {"$regex": search, "$options": "i"}},
        ]
    skip = (page - 1) * limit
    total = await db.equipment.count_documents(query)
    equipment = await db.equipment.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return {"equipment": [decrypt_equipment(e) for e in equipment], "total": total,
            "page": page, "pages": max(1, (total + limit - 1) // limit)}

# ==================== AUDITS ====================
@api_router.post("/audits")
async def create_audit(input: AuditCreateInput, user=Depends(get_current_user)):
    store = await db.stores.find_one({"cr_tienda": input.cr_tienda}, {"_id": 0})
    if not store:
        raise HTTPException(404, "Store not found")
    # Check for active audit (in_progress) OR pending_photos (waiting for photos)
    active = await db.audits.find_one(
        {"cr_tienda": input.cr_tienda, "status": {"$in": ["in_progress", "pending_photos"]}},
        {"_id": 0}
    )
    if active:
        # Ensure store is marked correctly
        await db.stores.update_one({"cr_tienda": input.cr_tienda}, {"$set": {
            "audit_status": active["status"], "last_audit_id": active["id"]
        }})
        # Block other users from entering an in-progress audit (pending_photos always accessible to owner)
        is_owner = active.get("auditor_id") == user["sub"]
        is_super  = user["perfil"] == "Super Administrador"
        if active["status"] == "in_progress" and not is_owner and not is_super:
            raise HTTPException(403, {
                "code": "AUDIT_IN_PROGRESS",
                "auditor_name": active.get("auditor_name", "otro usuario"),
                "started_at": active.get("started_at"),
                "tienda": active.get("tienda"),
            })
        return active
    audit = {
        "id": str(uuid.uuid4()), "cr_tienda": input.cr_tienda, "tienda": store["tienda"],
        "plaza": store["plaza"], "cr_plaza": store["cr_plaza"],
        "auditor_id": user["sub"], "auditor_name": user["nombre"],
        "started_at": datetime.now(timezone.utc).isoformat(), "finished_at": None,
        "status": "in_progress", "located_count": 0, "surplus_count": 0,
        "not_found_count": 0, "not_found_value": 0, "total_equipment": store["total_equipment"],
        "notes": ""
    }
    await db.audits.insert_one(audit)
    # Mark store as audit in-progress
    await db.stores.update_one({"cr_tienda": input.cr_tienda}, {"$set": {"audit_status": "in_progress", "last_audit_id": audit["id"]}})
    return {k: v for k, v in audit.items() if k != "_id"}


@api_router.get("/audits/stats/summary")
async def audits_stats_summary(plaza: Optional[str] = None, user=Depends(get_current_user)):
    """Stats: total audits done and stores audited. Optionally filtered by plaza."""
    audit_query: dict = {"status": {"$in": ["completed", "incompleto"]}}
    store_query: dict = {}
    if plaza:
        audit_query["plaza"] = plaza
        store_query["plaza"] = plaza
    total_audits = await db.audits.count_documents(audit_query)
    stores_audited = await db.audits.distinct("cr_tienda", audit_query)
    total_stores = await db.stores.count_documents(store_query)
    return {
        "total_audits": total_audits,
        "stores_audited": len(stores_audited),
        "total_stores": total_stores,
    }

@api_router.get("/audits/{audit_id}")
async def get_audit(audit_id: str, user=Depends(get_current_user)):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    return audit

async def _cancel_pending_baja(equipment_id: str, audit_id: str, cancelled_by: str, now_iso: str) -> Optional[dict]:
    """Cancel any pending BAJA movement for this equipment and return the cancelled movement."""
    pending_baja = await db.movements.find_one(
        {"equipment_id": equipment_id, "type": "baja", "status": "pending"}, {"_id": 0}
    )
    if pending_baja:
        await db.movements.update_one(
            {"id": pending_baja["id"]},
            {"$set": {
                "status": "cancelled", "cancelled_at": now_iso,
                "cancelled_by": cancelled_by,
                "cancel_reason": f"Equipo localizado en auditoria {audit_id} — BAJA revertida automaticamente"
            }}
        )
        return pending_baja
    return None

@api_router.post("/audits/{audit_id}/scan")
async def scan_barcode(audit_id: str, input: ScanInput, user=Depends(get_current_user)):
    barcode = input.barcode.strip().replace('\u202d', '').replace('\u202c', '')
    if not barcode:
        raise HTTPException(400, "Barcode required")

    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    if audit["status"] != "in_progress":
        raise HTTPException(400, "Audit already completed")

    existing = await db.audit_scans.find_one({"audit_id": audit_id, "codigo_barras": barcode}, {"_id": 0})
    if existing:
        return {"status": "already_scanned", "scan": existing}

    cr_tienda = audit["cr_tienda"]
    now_iso = datetime.now(timezone.utc).isoformat()

    # Check in current store
    equipment = await db.equipment.find_one({"cr_tienda": cr_tienda, "codigo_barras": barcode}, {"_id": 0})
    if equipment:
        cancelled_baja = await _cancel_pending_baja(equipment["id"], audit_id, user["nombre"], now_iso)
        scan_rec = {
            "id": str(uuid.uuid4()), "audit_id": audit_id, "codigo_barras": barcode,
            "equipment_id": equipment["id"], "classification": "localizado",
            "equipment_data": equipment, "origin_store": None,
            "scanned_at": now_iso, "scanned_by": user["nombre"],
            "baja_revertida": cancelled_baja is not None,
        }
        await db.audit_scans.insert_one(scan_rec)
        await db.audits.update_one({"id": audit_id}, {"$inc": {"located_count": 1}})
        return {
            "status": "localizado",
            "scan": {k: v for k, v in scan_rec.items() if k != "_id"},
            "baja_revertida": cancelled_baja is not None,
        }

    # Check in other stores (equipment assigned to a different store)
    other = await db.equipment.find_one({"codigo_barras": barcode, "cr_tienda": {"$ne": cr_tienda}}, {"_id": 0})
    if other:
        cancelled_baja = await _cancel_pending_baja(other["id"], audit_id, user["nombre"], now_iso)
        scan_rec = {
            "id": str(uuid.uuid4()), "audit_id": audit_id, "codigo_barras": barcode,
            "equipment_id": other["id"], "classification": "sobrante",
            "equipment_data": other,
            "origin_store": {"cr_tienda": other["cr_tienda"], "tienda": other["tienda"], "plaza": other["plaza"]},
            "scanned_at": now_iso, "scanned_by": user["nombre"],
            "baja_revertida": cancelled_baja is not None,
        }
        await db.audit_scans.insert_one(scan_rec)
        await db.audits.update_one({"id": audit_id}, {"$inc": {"surplus_count": 1}})
        return {
            "status": "sobrante",
            "scan": {k: v for k, v in scan_rec.items() if k != "_id"},
            "baja_revertida": cancelled_baja is not None,
        }

    # Not found anywhere in equipment DB
    scan_rec = {
        "id": str(uuid.uuid4()), "audit_id": audit_id, "codigo_barras": barcode,
        "equipment_id": None, "classification": "sobrante_desconocido",
        "equipment_data": None, "origin_store": None,
        "scanned_at": now_iso, "scanned_by": user["nombre"]
    }
    await db.audit_scans.insert_one(scan_rec)
    await db.audits.update_one({"id": audit_id}, {"$inc": {"surplus_count": 1}})
    return {"status": "sobrante_desconocido", "scan": {k: v for k, v in scan_rec.items() if k != "_id"}}

@api_router.get("/audits/{audit_id}/scans")
async def get_audit_scans(audit_id: str, user=Depends(get_current_user)):
    scans = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).sort("scanned_at", -1).to_list(10000)
    return scans

@api_router.post("/audits/{audit_id}/finalize")
async def finalize_audit(audit_id: str, input: Optional[FinalizeWithPhotosInput] = None, user=Depends(get_current_user)):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    if audit["status"] != "in_progress":
        raise HTTPException(400, "Audit already completed")

    cr_tienda = audit["cr_tienda"]
    scans = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000)

    # ── Auto-cancel if no equipment was scanned ──────────────────────────────
    # Finalizing an audit with zero scans is equivalent to not auditing at all.
    # Cancel it automatically so no ghost audits remain in the system.
    if not scans:
        now_iso = datetime.now(timezone.utc).isoformat()
        await db.audits.update_one({"id": audit_id}, {"$set": {
            "status": "cancelada", "finished_at": now_iso,
            "cancel_reason": "Auto-cancelada: ningún equipo fue escaneado al finalizar"
        }})
        await db.stores.update_one({"cr_tienda": cr_tienda}, {"$set": {
            "audited": False, "last_audit_id": None, "audit_status": None
        }})
        await sec_log(SecEvent.AUDIT_CANCELLED, actor_email=user["email"], actor_id=user["sub"],
                      target=audit_id, detail={"tienda": audit.get("tienda"), "reason": "no_scans"})
        raise HTTPException(400, "No se escaneó ningún equipo. La auditoría fue cancelada automáticamente.")

    scanned_ids = set(s["equipment_id"] for s in scans if s["equipment_id"])
    scanned_barcodes = set(s["codigo_barras"] for s in scans)

    all_eq = await db.equipment.find({"cr_tienda": cr_tienda}, {"_id": 0}).to_list(10000)
    now_iso = datetime.now(timezone.utc).isoformat()

    not_found_items = []
    not_found_value = 0
    scan_docs = []
    movement_docs = []

    for eq in all_eq:
        if eq["id"] not in scanned_ids and eq["codigo_barras"] not in scanned_barcodes:
            not_found_items.append(eq)
            not_found_value += eq.get("valor_real", 0)
            scan_docs.append({
                "id": str(uuid.uuid4()), "audit_id": audit_id, "codigo_barras": eq["codigo_barras"],
                "equipment_id": eq["id"], "classification": "no_localizado",
                "equipment_data": eq, "origin_store": None,
                "scanned_at": now_iso, "scanned_by": "system"
            })
            movement_docs.append({
                "id": str(uuid.uuid4()), "audit_id": audit_id, "equipment_id": eq["id"],
                "type": "baja", "from_cr_tienda": cr_tienda, "to_cr_tienda": None,
                "status": "pending", "created_at": now_iso,
                "created_by": user["nombre"], "created_by_id": user["sub"],
                "equipment_data": eq, "from_tienda": audit["tienda"], "to_tienda": None,
                "plaza": audit.get("plaza", ""), "auto_generated": True
            })

    # Bulk insert for performance — avoids N sequential round-trips to MongoDB
    if scan_docs:
        await db.audit_scans.insert_many(scan_docs)
    if movement_docs:
        await db.movements.insert_many(movement_docs)

    not_found_count = len(not_found_items)
    total_eq_count = len(all_eq)

    # ── Check if photos are needed and not yet provided ──────────────────
    settings = await db.system_settings.find_one({"_id": "global"}) or {}
    photo_required_alta  = settings.get("photo_required_alta", True)
    photo_required_baja  = settings.get("photo_required_baja", True)
    photo_required_ab    = photo_required_alta or photo_required_baja  # AB format needed if either is on
    photo_required_transf = settings.get("photo_required_transf", True)
    ttl_hours = int(settings.get("pending_photos_ttl_hours", 24))

    # ── Determine which manual movements exist for this audit ────────────────
    # Rules:
    #   • ALTA movements only come from sobrante_desconocido registration → always require photo if setting is ON
    #   • BAJA/disposal manual movements = auditor explicitly disposed equipment → require photo if setting ON
    #   • TRANSFER movements → require photo if setting ON
    #   • Auto-generated BAJA movements (no_localizado) = created by finalize itself; they may or may not
    #     have the auto_generated flag (older docs may lack it). We identify them by:
    #     auto_generated=True OR (type="baja" AND status="pending" AND created_by_id=system)
    #     The safest approach: ALTA movements are NEVER auto-generated (only manual sobrante creates them).
    #     BAJA movements can be auto-generated. So we use type-specific logic:
    #       - has_alta  = any movement with type="alta" for this audit  (always manual)
    #       - has_baja  = any movement with type in ("baja","disposal") AND auto_generated IS NOT True
    #       - has_transf = any movement with type="transfer" (always manual)
    movements_check = await db.movements.find(
        {"audit_id": audit_id},
        {"type": 1, "auto_generated": 1, "status": 1}
    ).to_list(1000)

    # ALTA movements are ALWAYS manual — sobrante_desconocido workflow only
    has_alta_movements = any(m["type"] == "alta" for m in movements_check)
    # BAJA/disposal: exclude auto_generated=True; also exclude cancelled ones (won't request photo)
    has_baja_movements = any(
        m["type"] in ("baja", "disposal")
        and not m.get("auto_generated")
        and m.get("status") != "cancelled"
        for m in movements_check
    )
    # TRANSFER movements are always manual
    has_transf_movements = any(m["type"] == "transfer" for m in movements_check)

    needs_photo_ab    = (has_alta_movements and photo_required_alta) or (has_baja_movements and photo_required_baja)
    needs_photo_transf = has_transf_movements and photo_required_transf

    # Photos provided inline in this request
    photos = {}
    if input:
        if input.photo_ab_base64:
            photos["photo_ab"] = input.photo_ab_base64
        if input.photo_transfer_base64:
            photos["photo_transf"] = input.photo_transfer_base64

    # If photos are required but not all provided → go to pending_photos state
    photos_satisfied = (
        (not needs_photo_ab    or "photo_ab"    in photos) and
        (not needs_photo_transf or "photo_transf" in photos)
    )
    if (needs_photo_ab or needs_photo_transf) and not photos_satisfied:
        final_status = "pending_photos"
        deadline_iso = (datetime.now(timezone.utc) + timedelta(hours=ttl_hours)).isoformat()
    else:
        final_status = "completed"
        deadline_iso = None

    update_data = {
        "status": final_status, "finished_at": now_iso,
        "not_found_count": not_found_count, "not_found_value": round(not_found_value, 2),
        "needs_photo_ab": needs_photo_ab, "needs_photo_transf": needs_photo_transf,
    }
    if deadline_iso:
        update_data["photos_deadline"] = deadline_iso
    if photos:
        update_data.update(photos)

    await db.audits.update_one({"id": audit_id}, {"$set": update_data})
    await db.stores.update_one({"cr_tienda": cr_tienda}, {"$set": {
        "audited": True, "last_audit_date": now_iso, "last_audit_id": audit_id,
        "audit_status": final_status
    }})

    updated_audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    all_scans = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000)
    movements = await db.movements.find({"audit_id": audit_id}, {"_id": 0}).to_list(1000)

    # ── Digital Signature ────────────────────────────────────────────────
    signature = sign_audit(updated_audit, all_scans)
    await db.audits.update_one({"id": audit_id}, {"$set": {"signature": signature}})
    updated_audit["signature"] = signature

    # Security audit log
    await sec_log(SecEvent.AUDIT_FINALIZED,
                  actor_email=user["email"], actor_id=user["sub"],
                  target=audit_id,
                  detail={"tienda": updated_audit.get("tienda"), "plaza": updated_audit.get("plaza"),
                           "status": final_status, "not_found": not_found_count,
                           "signature": signature["hash"][:16] + "…"})

    return {
        "audit": updated_audit,
        "summary": {
            "total_equipment": len(all_eq),
            "located": len([s for s in all_scans if s["classification"] == "localizado"]),
            "surplus": len([s for s in all_scans if s["classification"] in ("sobrante", "sobrante_desconocido")]),
            "not_found": len(not_found_items), "not_found_value": round(not_found_value, 2),
            "not_found_deprecated": len([e for e in not_found_items if e.get("depreciado", False)])
        },
        "scans": all_scans, "not_found_items": not_found_items, "movements": movements
    }

@api_router.get("/audits/{audit_id}/verify-signature")
async def verify_audit_sig(audit_id: str, user=Depends(get_current_user)):
    """Verify the HMAC-SHA256 digital signature of a finalized audit."""
    audit, scans = await asyncio.gather(
        db.audits.find_one({"id": audit_id}, {"_id": 0}),
        db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000),
    )
    if not audit:
        raise HTTPException(404, "Auditoría no encontrada")
    if audit.get("status") == "in_progress":
        raise HTTPException(400, "La auditoría aún no ha sido finalizada")
    if not audit.get("signature"):
        await sec_log(SecEvent.SIGNATURE_INVALID, actor_email=user["email"], actor_id=user["sub"],
                      target=audit_id, detail={"reason": "no_signature"})
        return {"valid": False, "reason": "Auditoría sin firma digital (fue finalizada antes de esta función)"}
    valid = verify_audit_signature(audit, scans)
    event = SecEvent.SIGNATURE_VALID if valid else SecEvent.SIGNATURE_INVALID
    await sec_log(event, actor_email=user["email"], actor_id=user["sub"],
                  target=audit_id, detail={"tienda": audit.get("tienda"), "valid": valid})
    return {
        "valid": valid,
        "signature": audit.get("signature"),
        "reason": None if valid else "Los datos de la auditoría han sido modificados después de su firma"
    }

@api_router.post("/audits/{audit_id}/cancel")
async def cancel_audit(audit_id: str, input: AuditCancelInput, user=Depends(get_current_user)):
    if not input.reason or not input.reason.strip():
        raise HTTPException(400, "Se requiere el motivo de cancelación")
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Auditoría no encontrada")
    if audit["status"] != "in_progress":
        raise HTTPException(400, "Solo se pueden cancelar auditorías en progreso")
    now_iso = datetime.now(timezone.utc).isoformat()
    await db.audits.update_one({"id": audit_id}, {"$set": {
        "status": "cancelada", "finished_at": now_iso,
        "cancel_reason": input.reason.strip(), "cancelled_by": user["nombre"]
    }})
    # Release store from audit
    await db.stores.update_one({"cr_tienda": audit["cr_tienda"]}, {"$set": {
        "audited": False, "last_audit_id": None, "audit_status": None
    }})
    await sec_log(SecEvent.AUDIT_CANCELLED, actor_email=user["email"], actor_id=user["sub"],
                  target=audit_id, detail={"tienda": audit.get("tienda"), "reason": input.reason.strip()})
    return {"message": "Auditoría cancelada", "cancel_reason": input.reason.strip()}

@api_router.post("/audits/{audit_id}/register-unknown-surplus")
async def register_unknown_surplus(audit_id: str, input: UnknownSurplusInput, user=Depends(get_current_user)):
    try:
        audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
        if not audit:
            raise HTTPException(status_code=404, detail="Auditoría no encontrada")
        if audit["status"] != "in_progress":
            raise HTTPException(status_code=400, detail="Solo se puede registrar en auditorías activas")

        barcode = input.codigo_barras.strip()
        cr_tienda = audit["cr_tienda"]
        now_iso = datetime.now(timezone.utc).isoformat()

        # Check if already registered in a previous attempt (idempotent)
        existing_eq = await db.equipment.find_one(
            {"cr_tienda": cr_tienda, "codigo_barras": barcode, "alta_manual": True},
            {"_id": 0}
        )
        if existing_eq:
            # Already inserted — just ensure scan and movement are updated
            new_eq_clean = existing_eq
        else:
            # Build equipment dict BEFORE insert to avoid _id contamination
            new_eq_clean = {
                "id": str(uuid.uuid4()),
                "cr_plaza": audit.get("cr_plaza", ""),
                "plaza": audit.get("plaza", ""),
                "cr_tienda": cr_tienda,
                "tienda": audit["tienda"],
                "codigo_barras": barcode,
                "no_activo": (input.no_activo or "").strip(),
                "mes_adquisicion": datetime.now().month,
                "año_adquisicion": datetime.now().year,
                "factura": "", "costo": 0.0, "depreciacion": 0.0,
                "vida_util": 0, "remanente": 0,
                "descripcion": input.descripcion.strip(),
                "marca": input.marca.strip(),
                "modelo": input.modelo.strip(),
                "serie": (input.serie or "").strip(),
                "meses_transcurridos": 0, "vida_util_restante": 0,
                "valor_real": 0.0, "depreciado": False, "alta_manual": True,
                "registered_at": now_iso, "registered_by": user["nombre"]
            }
            # Insert a COPY so insert_one does not pollute new_eq_clean with _id
            await db.equipment.insert_one(dict(new_eq_clean))
            await db.stores.update_one({"cr_tienda": cr_tienda}, {"$inc": {"total_equipment": 1}})

        # Update scan — find by audit+barcode only (classification may have changed on retry)
        await db.audit_scans.update_one(
            {"audit_id": audit_id, "codigo_barras": barcode},
            {"$set": {
                "equipment_id": new_eq_clean["id"],
                "equipment_data": new_eq_clean,
                "registered_manually": True,
                "classification": "sobrante_desconocido"
            }}
        )

        # Create ALTA movement only if one doesn't exist yet
        existing_mov = await db.movements.find_one(
            {"audit_id": audit_id, "equipment_id": new_eq_clean["id"], "type": "alta"},
            {"_id": 0}
        )
        if not existing_mov:
            alta_movement = {
                "id": str(uuid.uuid4()),
                "audit_id": audit_id,
                "equipment_id": new_eq_clean["id"],
                "type": "alta",
                "from_cr_tienda": None,
                "to_cr_tienda": cr_tienda,
                "status": "pending",
                "created_at": now_iso,
                "created_by": user["nombre"],
                "created_by_id": user["sub"],
                "equipment_data": new_eq_clean,
                "from_tienda": None,
                "to_tienda": audit["tienda"],
                "plaza": audit.get("plaza", "")
            }
            await db.movements.insert_one(dict(alta_movement))
        else:
            alta_movement = existing_mov

        return {
            "message": "Equipo registrado como ALTA",
            "equipment": new_eq_clean,
            "movement": {k: v for k, v in alta_movement.items() if k != "_id"}
        }
    except HTTPException:
        raise
    except Exception as e:
        logger.error(f"Error en register-unknown-surplus: {e}", exc_info=True)
        raise HTTPException(status_code=500, detail=f"Error interno al registrar el equipo: {str(e)}")

@api_router.get("/audits/{audit_id}/summary")
async def get_audit_summary(audit_id: str, user=Depends(get_current_user)):
    # Fetch audit + scans + movements in parallel
    audit, scans, movements = await asyncio.gather(
        db.audits.find_one({"id": audit_id}, {"_id": 0}),
        db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000),
        db.movements.find({"audit_id": audit_id}, {"_id": 0}).to_list(1000),
    )
    if not audit:
        raise HTTPException(404, "Audit not found")
    located    = [s for s in scans if s["classification"] == "localizado"]
    surplus    = [s for s in scans if s["classification"] in ("sobrante", "sobrante_desconocido")]
    not_found  = [s for s in scans if s["classification"] == "no_localizado"]
    nf_value   = sum((s.get("equipment_data") or {}).get("valor_real", 0) for s in not_found)
    nf_dep     = len([s for s in not_found if (s.get("equipment_data") or {}).get("depreciado", False)])
    return {
        "audit": audit, "located": located, "surplus": surplus, "not_found": not_found, "movements": movements,
        "stats": {"total_equipment": audit.get("total_equipment", 0), "located_count": len(located),
                  "surplus_count": len(surplus), "not_found_count": len(not_found),
                  "not_found_value": round(nf_value, 2), "not_found_deprecated": nf_dep, "movements_count": len(movements)}
    }

# ==================== DELETE AUDIT (Super Admin) ====================
@api_router.delete("/audits/{audit_id}")
async def delete_audit(audit_id: str, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    # Snapshot BEFORE delete for rollback
    scans_snap = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000)
    movements_snap = await db.movements.find({"audit_id": audit_id}, {"_id": 0}).to_list(1000)
    await save_history(HistAction.DELETE_AUDIT, user["email"], user["sub"],
                       audit_id, f"{audit.get('tienda')} ({audit.get('cr_tienda')})",
                       dict(audit), {"scans": scans_snap, "movements": movements_snap})
    # Delete associated scans and movements
    await db.audit_scans.delete_many({"audit_id": audit_id})
    await db.movements.delete_many({"audit_id": audit_id})
    # Reset store audit status so it can be audited again
    cr_tienda = audit.get("cr_tienda")
    if cr_tienda:
        await db.stores.update_one({"cr_tienda": cr_tienda}, {"$set": {
            "audited": False, "last_audit_date": None, "last_audit_id": None, "audit_status": None
        }})
    await db.audits.delete_one({"id": audit_id})
    await sec_log(SecEvent.AUDIT_DELETED, actor_email=user["email"], actor_id=user["sub"],
                  target=audit_id, detail={"tienda": audit.get("tienda"), "plaza": audit.get("plaza"),
                                            "status": audit.get("status")})
    return {"message": "Audit deleted", "cr_tienda": cr_tienda}

# ==================== MOVEMENTS ====================

class NotesInput(BaseModel):
    notes: str

class MovementInputExtended(BaseModel):
    audit_id: str
    equipment_id: Optional[str] = None
    type: str
    from_cr_tienda: Optional[str] = None
    to_cr_tienda: Optional[str] = None
    extra_data: Optional[dict] = None

@api_router.post("/audits/{audit_id}/photos")
async def upload_audit_photos(audit_id: str, photo_ab: Optional[UploadFile] = File(None), photo_transf: Optional[UploadFile] = File(None), user=Depends(get_current_user)):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Audit not found")
    update = {}
    if photo_ab:
        content = await photo_ab.read()
        update["photo_ab"] = base64.b64encode(content).decode()
        update["photo_ab_filename"] = photo_ab.filename
    if photo_transf:
        content = await photo_transf.read()
        update["photo_transf"] = base64.b64encode(content).decode()
        update["photo_transf_filename"] = photo_transf.filename
    if update:
        update["photos_uploaded_at"] = datetime.now(timezone.utc).isoformat()
        update["photos_uploaded_by"] = user["nombre"]
        await db.audits.update_one({"id": audit_id}, {"$set": update})

        # If all required photos are now present, transition pending_photos → completed
        if audit.get("status") == "pending_photos":
            updated = await db.audits.find_one({"id": audit_id}, {"_id": 0})
            needs_ab    = updated.get("needs_photo_ab", False)
            needs_transf = updated.get("needs_photo_transf", False)
            has_ab      = bool(updated.get("photo_ab"))
            has_transf  = bool(updated.get("photo_transf"))
            if (not needs_ab or has_ab) and (not needs_transf or has_transf):
                now_iso = datetime.now(timezone.utc).isoformat()
                all_scans = await db.audit_scans.find({"audit_id": audit_id}, {"_id": 0}).to_list(10000)
                signature = sign_audit(updated, all_scans)
                await db.audits.update_one({"id": audit_id}, {"$set": {
                    "status": "completed", "photos_deadline": None, "signature": signature
                }})
                await db.stores.update_one({"cr_tienda": audit.get("cr_tienda")}, {"$set": {
                    "audit_status": "completed"
                }})
                return {"message": "Fotos guardadas y auditoría completada", "completed": True}

    return {"message": "Photos saved", "completed": False}

@api_router.put("/audits/{audit_id}/notes")
async def update_audit_notes(audit_id: str, input: NotesInput, user=Depends(get_current_user)):
    result = await db.audits.update_one({"id": audit_id}, {"$set": {"notes": input.notes}})
    if result.matched_count == 0:
        raise HTTPException(404, "Audit not found")
    return {"message": "Notes updated"}

@api_router.delete("/audits/{audit_id}/scans/{scan_id}")
async def delete_scan(audit_id: str, scan_id: str, user=Depends(get_current_user)):
    audit = await db.audits.find_one({"id": audit_id}, {"_id": 0})
    if not audit or audit["status"] != "in_progress":
        raise HTTPException(400, "Cannot delete scan")
    scan = await db.audit_scans.find_one({"id": scan_id, "audit_id": audit_id}, {"_id": 0})
    if not scan:
        raise HTTPException(404, "Scan not found")
    # Update audit counts
    dec = {}
    if scan["classification"] == "localizado":
        dec["located_count"] = -1
    elif scan["classification"] in ("sobrante", "sobrante_desconocido"):
        dec["surplus_count"] = -1
    if dec:
        await db.audits.update_one({"id": audit_id}, {"$inc": dec})
    await db.audit_scans.delete_one({"id": scan_id})
    return {"message": "Scan deleted"}

@api_router.post("/movements")
async def create_movement(input: MovementInputExtended, user=Depends(get_current_user)):
    movement = {
        "id": str(uuid.uuid4()), "audit_id": input.audit_id, "equipment_id": input.equipment_id,
        "type": input.type, "from_cr_tienda": input.from_cr_tienda, "to_cr_tienda": input.to_cr_tienda,
        "status": "pending", "created_at": datetime.now(timezone.utc).isoformat(),
        "created_by": user["nombre"], "created_by_id": user["sub"]
    }
    if input.extra_data:
        movement["extra_data"] = input.extra_data
    if input.equipment_id:
        eq = await db.equipment.find_one({"id": input.equipment_id}, {"_id": 0})
        if eq:
            movement["equipment_data"] = eq
    if input.from_cr_tienda:
        fs = await db.stores.find_one({"cr_tienda": input.from_cr_tienda}, {"_id": 0})
        movement["from_tienda"] = fs["tienda"] if fs else ""
    if input.to_cr_tienda:
        ts = await db.stores.find_one({"cr_tienda": input.to_cr_tienda}, {"_id": 0})
        movement["to_tienda"] = ts["tienda"] if ts else ""
    await db.movements.insert_one(movement)
    return {k: v for k, v in movement.items() if k != "_id"}

@api_router.get("/movements")
async def get_movements(audit_id: Optional[str] = None, type: Optional[str] = None, page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    query = {}
    if audit_id:
        query["audit_id"] = audit_id
    if type:
        query["type"] = type
    skip = (page - 1) * limit
    total = await db.movements.count_documents(query)
    items = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"movements": items, "total": total, "page": page}

# ==================== LOGS ====================
@api_router.get("/logs/classifications")
async def get_classification_logs(audit_id: Optional[str] = None, classification: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 500, user=Depends(get_current_user)):
    query = {}
    if audit_id:
        query["audit_id"] = audit_id
    if classification:
        query["classification"] = classification
    if search:
        query["$or"] = [
            {"codigo_barras": {"$regex": search, "$options": "i"}},
            {"equipment_data.tienda": {"$regex": search, "$options": "i"}},
            {"equipment_data.cr_tienda": {"$regex": search, "$options": "i"}},
        ]
    skip = (page - 1) * limit
    total = await db.audit_scans.count_documents(query)
    items = await db.audit_scans.find(query, {"_id": 0}).sort("scanned_at", -1).skip(skip).limit(limit).to_list(limit)
    return {"items": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.get("/logs/movements")
async def get_movement_logs(type: Optional[str] = None, search: Optional[str] = None, page: int = 1, limit: int = 100, user=Depends(get_current_user)):
    query = {}
    if type and type != "all":
        if type == "bajas":
            query["type"] = {"$in": ["baja", "disposal"]}
        elif type == "altas":
            query["type"] = "alta"
        elif type == "transferencias":
            query["type"] = "transfer"
        else:
            query["type"] = type
    if search:
        query["$or"] = [
            {"from_tienda": {"$regex": search, "$options": "i"}},
            {"to_tienda": {"$regex": search, "$options": "i"}},
            {"from_cr_tienda": {"$regex": search, "$options": "i"}},
            {"to_cr_tienda": {"$regex": search, "$options": "i"}},
            {"equipment_data.codigo_barras": {"$regex": search, "$options": "i"}},
        ]
    skip = (page - 1) * limit
    total = await db.movements.count_documents(query)
    items = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).skip(skip).limit(limit).to_list(limit)
    mv_query = dict(query)
    tv_q = {**mv_query, "type": "transfer"}
    dv_q = {**mv_query, "type": {"$in": ["disposal", "baja"]}}
    av_q = {**mv_query, "type": "alta"}
    tv = await db.movements.aggregate([{"$match": tv_q}, {"$group": {"_id": None, "t": {"$sum": "$equipment_data.valor_real"}}}]).to_list(1)
    dv = await db.movements.aggregate([{"$match": dv_q}, {"$group": {"_id": None, "t": {"$sum": "$equipment_data.valor_real"}}}]).to_list(1)
    av = await db.movements.aggregate([{"$match": av_q}, {"$group": {"_id": None, "t": {"$sum": "$equipment_data.valor_real"}}}]).to_list(1)
    return {"items": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit),
            "transfer_total_value": round(tv[0]["t"], 2) if tv else 0,
            "disposal_total_value": round(dv[0]["t"], 2) if dv else 0,
            "alta_total_value": round(av[0]["t"], 2) if av else 0}

@api_router.get("/logs/audits")
async def get_audit_logs(status: Optional[str] = None, search: Optional[str] = None, plaza: Optional[str] = None,
                         sort_by: str = "started_at", sort_dir: str = "desc",
                         page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    query = {}
    if status:
        query["status"] = status
    if plaza:
        query["plaza"] = plaza
    if search:
        query["$or"] = [
            {"tienda": {"$regex": search, "$options": "i"}},
            {"cr_tienda": {"$regex": search, "$options": "i"}},
            {"plaza": {"$regex": search, "$options": "i"}},
        ]
    skip = (page - 1) * limit
    # Allowed sort fields — whitelist to prevent injection
    ALLOWED_SORT = {"started_at", "finished_at", "cr_tienda", "tienda", "plaza",
                    "auditor_name", "status", "located_count", "not_found_count"}
    sort_field = sort_by if sort_by in ALLOWED_SORT else "started_at"
    sort_direction = -1 if sort_dir == "desc" else 1
    _AUDIT_LIST_PROJECTION = {"_id": 0, "photo_ab": 0, "photo_transf": 0}
    total, items = await asyncio.gather(
        db.audits.count_documents(query),
        db.audits.find(query, _AUDIT_LIST_PROJECTION).sort(sort_field, sort_direction).skip(skip).limit(limit).to_list(limit)
    )
    return {"items": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

# ==================== PDF DOWNLOADS ====================
@api_router.get("/download/manual")
async def download_manual(user=Depends(get_current_user)):
    stats = None
    try:
        total_stores = await db.stores.count_documents({})
        audited_stores = await db.stores.count_documents({"audited": True})
        total_equipment = await db.equipment.count_documents({})
        deprecated_eq = await db.equipment.count_documents({"depreciado": True})
        val_pipe = [{"$group": {"_id": None, "total_cost": {"$sum": "$costo"}, "total_real": {"$sum": "$valor_real"}}}]
        val = await db.equipment.aggregate(val_pipe).to_list(1)
        plaza_pipe = [{"$group": {"_id": "$plaza", "count": {"$sum": 1}}}]
        plaza_stats = await db.equipment.aggregate(plaza_pipe).to_list(100)
        stats = {
            "total_stores": total_stores, "audited_stores": audited_stores,
            "total_equipment": total_equipment, "deprecated_equipment": deprecated_eq,
            "active_equipment": total_equipment - deprecated_eq,
            "total_cost": val[0]["total_cost"] if val else 0,
            "total_real_value": val[0]["total_real"] if val else 0,
            "completed_audits": await db.audits.count_documents({"status": {"$in": ["completed"]}}),
            "equipment_by_plaza": {s["_id"]: s["count"] for s in plaza_stats if s["_id"]}
        }
    except Exception as e:
        logger.error(f"Error getting stats for manual: {e}")
    output = generate_user_manual(stats, [], user_perfil=user.get("perfil", "Administrador"))
    filename = "SIGAF_Manual_de_Usuario.pdf"
    if user.get("perfil") == "Socio Tecnologico":
        filename = "SIGAF_Manual_Socio_Tecnologico.pdf"
    elif user.get("perfil") == "Administrador":
        filename = "SIGAF_Manual_Administrador.pdf"
    return StreamingResponse(output, media_type="application/pdf",
        headers={"Content-Disposition": f"attachment; filename={filename}"})

@api_router.get("/download/presentation")
async def download_presentation(user=Depends(get_current_user)):
    stats = None
    try:
        total_stores = await db.stores.count_documents({})
        audited_stores = await db.stores.count_documents({"audited": True})
        total_equipment = await db.equipment.count_documents({})
        deprecated_eq = await db.equipment.count_documents({"depreciado": True})
        val_pipe = [{"$group": {"_id": None, "total_cost": {"$sum": "$costo"}, "total_real": {"$sum": "$valor_real"}}}]
        val = await db.equipment.aggregate(val_pipe).to_list(1)
        plaza_pipe = [{"$group": {"_id": "$plaza", "count": {"$sum": 1}}}]
        plaza_stats = await db.equipment.aggregate(plaza_pipe).to_list(100)
        stats = {
            "total_stores": total_stores, "audited_stores": audited_stores,
            "total_equipment": total_equipment, "deprecated_equipment": deprecated_eq,
            "active_equipment": total_equipment - deprecated_eq,
            "total_cost": val[0]["total_cost"] if val else 0,
            "total_real_value": val[0]["total_real"] if val else 0,
            "completed_audits": await db.audits.count_documents({"status": {"$in": ["completed"]}}),
            "equipment_by_plaza": {s["_id"]: s["count"] for s in plaza_stats if s["_id"]}
        }
    except Exception as e:
        logger.error(f"Error getting stats for presentation: {e}")
    output = generate_presentation(stats, [])
    return StreamingResponse(output, media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=SIGAF_Presentacion.pdf"})

# ==================== EXPORT ====================
async def get_user_from_token(token: str):
    try:
        return jwt.decode(token, SECRET_KEY, algorithms=["HS256"])
    except:
        return None

@api_router.get("/export/{export_type}")
async def export_to_excel(export_type: str, token: Optional[str] = None, authorization: Optional[str] = Header(None),
                          classification: Optional[str] = None, type: Optional[str] = None, status: Optional[str] = None,
                          search: Optional[str] = None, plaza: Optional[str] = None):
    # Accept token from query param OR header
    user = None
    if authorization and authorization.startswith("Bearer "):
        user = await get_user_from_token(authorization.split(" ")[1])
    if not user and token:
        user = await get_user_from_token(token)
    if not user:
        raise HTTPException(401, "Not authenticated")

    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    DARK_BLUE = "1E3C78"
    HEADER_BLUE = "2B5BA8"
    LIGHT_BLUE = "D6E4F7"
    WHITE = "FFFFFF"
    GRAY = "F5F5F5"

    def style_title_row(ws, row, text, merge_cols):
        ws.merge_cells(f"A{row}:{get_column_letter(merge_cols)}{row}")
        cell = ws.cell(row=row, column=1, value=text)
        cell.font = Font(bold=True, color=WHITE, size=14, name="Calibri")
        cell.fill = PatternFill(fill_type="solid", fgColor=DARK_BLUE)
        cell.alignment = Alignment(horizontal="center", vertical="center")
        ws.row_dimensions[row].height = 28

    def style_info_row(ws, row, label, value, merge_cols):
        ws.cell(row=row, column=1, value=label).font = Font(bold=True, size=10, name="Calibri")
        cell = ws.cell(row=row, column=2, value=value)
        cell.font = Font(size=10, name="Calibri")
        ws.row_dimensions[row].height = 18

    def style_header_row(ws, row, headers):
        for col, h in enumerate(headers, 1):
            cell = ws.cell(row=row, column=col, value=h)
            cell.font = Font(bold=True, color=WHITE, size=10, name="Calibri")
            cell.fill = PatternFill(fill_type="solid", fgColor=HEADER_BLUE)
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = Border(
                left=Side(style="thin", color=WHITE),
                right=Side(style="thin", color=WHITE),
                bottom=Side(style="medium", color=WHITE)
            )
        ws.row_dimensions[row].height = 20

    def style_data_row(ws, row, values, alt=False):
        fill_color = LIGHT_BLUE if alt else WHITE
        for col, val in enumerate(values, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill(fill_type="solid", fgColor=fill_color)
            cell.font = Font(size=9, name="Calibri")
            cell.border = Border(
                left=Side(style="thin", color="DDDDDD"),
                right=Side(style="thin", color="DDDDDD"),
                bottom=Side(style="thin", color="DDDDDD")
            )
            if isinstance(val, float) and col > 0:
                cell.number_format = '"$"#,##0.00'
            cell.alignment = Alignment(vertical="center")
        ws.row_dimensions[row].height = 16

    wb = openpyxl.Workbook()
    ws = wb.active
    today_str = datetime.now().strftime("%Y-%m-%d")
    today_display = datetime.now().strftime("%d/%m/%Y")
    plaza_name = plaza or "General"
    filename = f"sigaf_{export_type}_{today_str}.xlsx"

    if export_type == "classifications":
        ws.title = "Clasificaciones"
        query = {}
        if classification and classification != "all":
            query["classification"] = classification
        if search:
            query["$or"] = [
                {"codigo_barras": {"$regex": search, "$options": "i"}},
                {"equipment_data.tienda": {"$regex": search, "$options": "i"}},
            ]
        items = await db.audit_scans.find(query, {"_id": 0}).sort("scanned_at", -1).to_list(100000)

        style_title_row(ws, 1, "SIGAF - Reporte de Clasificaciones de Auditoría", 9)
        ws.cell(row=2, column=1, value=f"Fecha de exportación: {today_display}")
        ws.cell(row=2, column=1).font = Font(size=9, italic=True, name="Calibri", color="666666")
        ws.merge_cells("A2:I2")
        ws.row_dimensions[2].height = 15

        style_header_row(ws, 3, ["Fecha", "Código Barras", "Clasificación", "Descripción", "Marca", "Modelo", "Serie", "Depreciado", "Tienda"])
        for i, item in enumerate(items):
            eq = item.get("equipment_data") or {}
            dep = eq.get("depreciado", False)
            style_data_row(ws, i + 4, [
                item.get("scanned_at", "")[:19].replace("T", " ") if item.get("scanned_at") else "",
                item.get("codigo_barras", ""),
                item.get("classification", ""),
                eq.get("descripcion", ""),
                eq.get("marca", ""),
                eq.get("modelo", ""),
                eq.get("serie", ""),
                "Sí" if dep else "No",
                eq.get("tienda", ""),
            ], alt=(i % 2 == 0))

        for col, width in zip(range(1, 10), [18, 14, 16, 30, 14, 16, 18, 10, 24]):
            ws.column_dimensions[get_column_letter(col)].width = width

    elif export_type in ("movements-ab", "movements-transferencias", "movements"):
        # Determine movement types to export
        if export_type == "movements-ab":
            # Base query: altas, bajas, disposal
            base_types = ["alta", "baja", "disposal"]
            # Apply type filter if provided (from active filter in UI)
            if type == "bajas":
                query = {"type": {"$in": ["baja", "disposal"]}}
            elif type == "altas":
                query = {"type": "alta"}
            else:
                query = {"type": {"$in": base_types}}
            ws.title = "Altas y Bajas"
            doc_title = "Formato de Movimiento de AF — ALTAS y BAJAS"
            plaza_name = plaza or "General"
            filename = f"SIGAF_AB_{plaza_name}_{today_str}.xlsx"
            headers = ["Plaza Origen", "Tipo de Movimiento", "Número de Activo", "Código de Barras",
                       "Descripción del Equipo", "Valor Real", "Marca", "Modelo", "Año",
                       "Número de Serie", "CR Tienda Origen", "Tienda Origen"]
            merge_cols = 12
        elif export_type == "movements-transferencias":
            query = {"type": "transfer"}
            ws.title = "Transferencias"
            doc_title = "Formato de Movimiento de AF — TRANSFERENCIAS"
            plaza_name = plaza or "General"
            filename = f"SIGAF_TRANSFERENCIAS_{plaza_name}_{today_str}.xlsx"
            # Plaza Destino | Tipo | No Activo | Codigo Barras | Descripcion | Valor Real | Marca | Modelo | Año | Serie | CR Origen | Tienda Origen | CR Destino | Tienda Destino
            headers = ["Plaza Destino", "Tipo de Movimiento", "Número de Activo", "Código de Barras",
                       "Descripción del Equipo", "Valor Real", "Marca", "Modelo", "Año",
                       "Número de Serie", "CR Tienda Origen", "Tienda Origen",
                       "CR Tienda Destino", "Tienda Destino"]
            merge_cols = 14
        else:  # movements (all)
            query = {}
            if type and type != "all":
                query["type"] = type
            ws.title = "Movimientos"
            doc_title = "Formato de Movimiento de AF"
            filename = f"SIGAF_MOVIMIENTOS_{plaza_name}_{today_str}.xlsx"
            headers = ["Plaza", "Tipo de Movimiento", "Número de Activo", "Código de Barras",
                       "Descripción del Equipo", "Valor Real", "Marca", "Modelo", "Año",
                       "Número de Serie", "CR Tienda Origen", "Tienda Origen",
                       "CR Tienda Destino", "Tienda Destino"]
            merge_cols = 14

        if search:
            query["$or"] = [
                {"from_tienda": {"$regex": search, "$options": "i"}},
                {"to_tienda": {"$regex": search, "$options": "i"}},
                {"equipment_data.codigo_barras": {"$regex": search, "$options": "i"}},
            ]
        if plaza and plaza != "all":
            query["plaza"] = plaza

        items_raw = await db.movements.find(query, {"_id": 0}).sort("created_at", -1).to_list(100000)
        # Deduplicate: keep only the most recent movement per (equipment_id, type, audit_id).
        # Duplicates can occur when finalize is called multiple times (e.g. pending_photos flow).
        seen_keys = set()
        items = []
        for m in items_raw:
            dup_key = (m.get("equipment_id", ""), m.get("type", ""), m.get("audit_id", ""))
            if dup_key not in seen_keys:
                seen_keys.add(dup_key)
                items.append(m)

        # Lookup plazas for to_cr_tienda (needed for transferencias)
        store_plaza_cache = {}
        async def get_store_plaza(cr):
            if not cr:
                return ""
            if cr not in store_plaza_cache:
                s = await db.stores.find_one({"cr_tienda": cr}, {"_id": 0, "plaza": 1})
                store_plaza_cache[cr] = s.get("plaza", "") if s else ""
            return store_plaza_cache[cr]

        # Title
        style_title_row(ws, 1, doc_title, merge_cols)
        # Info rows — removed PLAZA from header as requested
        info_row = 2
        ws.cell(row=info_row, column=1, value="FECHA:").font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=info_row, column=2, value=today_display).font = Font(size=10, name="Calibri")
        ws.cell(row=info_row, column=4, value="DEPARTAMENTO:").font = Font(bold=True, size=10, name="Calibri")
        ws.cell(row=info_row, column=5, value="Sistemas").font = Font(size=10, name="Calibri")
        ws.row_dimensions[info_row].height = 20

        ws.row_dimensions[3].height = 5  # spacer row

        style_header_row(ws, 4, headers)

        tipo_labels = {"alta": "ALTA", "baja": "BAJA", "disposal": "BAJA", "transfer": "TRANSFERENCIA"}

        for i, item in enumerate(items):
            eq = item.get("equipment_data") or {}
            tipo = tipo_labels.get(item.get("type", ""), item.get("type", "").upper())
            from_cr = item.get("from_cr_tienda", "")
            to_cr = item.get("to_cr_tienda", "")

            # For AB: plaza of origin store. For transferencias: plaza of destination store.
            if export_type == "movements-ab":
                plaza_col = item.get("plaza", "") or eq.get("plaza", "")
            elif export_type == "movements-transferencias":
                plaza_col = await get_store_plaza(to_cr)
            else:
                plaza_col = item.get("plaza", "") or eq.get("plaza", "")

            row_vals = [
                plaza_col,
                tipo,
                eq.get("no_activo", ""),
                eq.get("codigo_barras", ""),
                eq.get("descripcion", ""),
                eq.get("valor_real", 0.0),
                eq.get("marca", ""),
                eq.get("modelo", ""),
                eq.get("año_adquisicion", ""),
                eq.get("serie", ""),
                from_cr,
                item.get("from_tienda", ""),
            ]
            if merge_cols >= 14:
                row_vals += [to_cr, item.get("to_tienda", "")]
            style_data_row(ws, i + 5, row_vals, alt=(i % 2 == 0))

        col_widths = [16, 16, 14, 14, 34, 12, 14, 16, 8, 18, 14, 26, 14, 26]
        for col, width in enumerate(col_widths[:merge_cols], 1):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Agregar hoja de imágenes de formatos de movimiento ──
        # Recolectar audit_ids únicos de los movimientos exportados
        if export_type in ("movements-ab", "movements-transferencias"):
            audit_ids = list({item.get("audit_id") for item in items if item.get("audit_id")})
            if audit_ids:
                photo_field = "photo_ab" if export_type == "movements-ab" else "photo_transf"
                photo_label = "Formato ALTAS/BAJAS" if export_type == "movements-ab" else "Formato TRANSFERENCIAS"
                audits_with_photos = await db.audits.find(
                    {"id": {"$in": audit_ids}, photo_field: {"$exists": True, "$ne": None}},
                    {"_id": 0, "id": 1, "tienda": 1, "cr_tienda": 1, "finished_at": 1, photo_field: 1}
                ).to_list(1000)
                if audits_with_photos:
                    ws_photos = wb.create_sheet(title="Imágenes Formatos")
                    ws_photos.column_dimensions["A"].width = 20
                    ws_photos.column_dimensions["B"].width = 20
                    ws_photos.column_dimensions["C"].width = 20
                    ws_photos.column_dimensions["D"].width = 80
                    style_title_row(ws_photos, 1, f"Imágenes — {photo_label}", 4)
                    style_header_row(ws_photos, 2, ["CR Tienda", "Tienda", "Fecha", photo_label])
                    img_row = 3
                    for audit_doc in audits_with_photos:
                        photo_b64 = audit_doc.get(photo_field)
                        tienda_name = audit_doc.get("tienda", "")
                        cr = audit_doc.get("cr_tienda", "")
                        fecha = (audit_doc.get("finished_at") or "")[:10]
                        ws_photos.cell(row=img_row, column=1, value=cr)
                        ws_photos.cell(row=img_row, column=2, value=tienda_name)
                        ws_photos.cell(row=img_row, column=3, value=fecha)
                        try:
                            import base64 as _b64i, io as _ioi
                            from openpyxl.drawing.image import Image as XLImage
                            img_bytes = _b64i.b64decode(photo_b64)
                            img_stream = _ioi.BytesIO(img_bytes)
                            xl_img = XLImage(img_stream)
                            # Escalar imagen a máximo 400x300 manteniendo proporción
                            max_w, max_h = 400, 300
                            if xl_img.width > max_w or xl_img.height > max_h:
                                ratio = min(max_w / xl_img.width, max_h / xl_img.height)
                                xl_img.width = int(xl_img.width * ratio)
                                xl_img.height = int(xl_img.height * ratio)
                            cell_ref = f"D{img_row}"
                            ws_photos.add_image(xl_img, cell_ref)
                            ws_photos.row_dimensions[img_row].height = xl_img.height * 0.75 + 10
                        except Exception:
                            ws_photos.cell(row=img_row, column=4, value="[Imagen no disponible]")
                        img_row += 1

    elif export_type == "audits":
        ws.title = "Auditorías"
        query = {}
        if status and status != "all":
            query["status"] = status
        if search:
            query["$or"] = [
                {"tienda": {"$regex": search, "$options": "i"}},
                {"cr_tienda": {"$regex": search, "$options": "i"}},
            ]
        items = await db.audits.find(query, {"_id": 0}).sort("started_at", -1).to_list(100000)

        style_title_row(ws, 1, "SIGAF - Historial de Auditorías", 13)
        ws.cell(row=2, column=1, value=f"Fecha de exportación: {today_display}")
        ws.cell(row=2, column=1).font = Font(size=9, italic=True, name="Calibri", color="666666")
        ws.merge_cells("A2:M2")
        ws.row_dimensions[2].height = 15

        style_header_row(ws, 3, ["Fecha Inicio", "Fecha Fin", "Tienda", "CR", "Plaza", "Auditor",
                                   "Estado", "Total Equipos", "Localizados", "Sobrantes",
                                   "No Localizados", "Valor No Localizado", "Motivo Cancelación"])
        for i, item in enumerate(items):
            style_data_row(ws, i + 4, [
                item.get("started_at", "")[:19].replace("T", " ") if item.get("started_at") else "",
                item.get("finished_at", "")[:19].replace("T", " ") if item.get("finished_at") else "",
                item.get("tienda", ""), item.get("cr_tienda", ""), item.get("plaza", ""),
                item.get("auditor_name", ""), item.get("status", ""),
                item.get("total_equipment", 0), item.get("located_count", 0), item.get("surplus_count", 0),
                item.get("not_found_count", 0), float(item.get("not_found_value", 0)),
                item.get("cancel_reason", ""),
            ], alt=(i % 2 == 0))

        for col, width in zip(range(1, 14), [18, 18, 28, 10, 14, 20, 12, 12, 12, 12, 14, 16, 28]):
            ws.column_dimensions[get_column_letter(col)].width = width

        # ── Hoja de imágenes de formatos (solo auditorías completadas) ──
        completed_with_photos = [
            item for item in items
            if item.get("status") == "completed" and (item.get("photo_ab") or item.get("photo_transf"))
        ]
        if completed_with_photos:
            ws_photos = wb.create_sheet(title="Imágenes Formatos")
            ws_photos.column_dimensions["A"].width = 20
            ws_photos.column_dimensions["B"].width = 28
            ws_photos.column_dimensions["C"].width = 18
            ws_photos.column_dimensions["D"].width = 70
            ws_photos.column_dimensions["E"].width = 70
            style_title_row(ws_photos, 1, "SIGAF — Formatos de Movimiento de Auditorías", 5)
            style_header_row(ws_photos, 2, ["CR Tienda", "Tienda", "Fecha", "Formato ALTAS/BAJAS", "Formato TRANSFERENCIAS"])
            img_row = 3
            for audit_doc in completed_with_photos:
                ws_photos.cell(row=img_row, column=1, value=audit_doc.get("cr_tienda", ""))
                ws_photos.cell(row=img_row, column=2, value=audit_doc.get("tienda", ""))
                fecha = (audit_doc.get("finished_at") or "")[:10]
                ws_photos.cell(row=img_row, column=3, value=fecha)
                row_height = 20
                for col_idx, field in [(4, "photo_ab"), (5, "photo_transf")]:
                    photo_b64 = audit_doc.get(field)
                    if photo_b64:
                        try:
                            import base64 as _b64i, io as _ioi
                            from openpyxl.drawing.image import Image as XLImage
                            img_bytes = _b64i.b64decode(photo_b64)
                            xl_img = XLImage(_ioi.BytesIO(img_bytes))
                            max_w, max_h = 380, 280
                            if xl_img.width > max_w or xl_img.height > max_h:
                                ratio = min(max_w / xl_img.width, max_h / xl_img.height)
                                xl_img.width = int(xl_img.width * ratio)
                                xl_img.height = int(xl_img.height * ratio)
                            cell_ref = f"{get_column_letter(col_idx)}{img_row}"
                            ws_photos.add_image(xl_img, cell_ref)
                            row_height = max(row_height, int(xl_img.height * 0.75) + 10)
                        except Exception:
                            ws_photos.cell(row=img_row, column=col_idx, value="[Imagen no disponible]")
                ws_photos.row_dimensions[img_row].height = row_height
                img_row += 1
    else:
        raise HTTPException(400, "Invalid export type")

    # Footer row
    last_row = ws.max_row + 2
    ws.cell(row=last_row, column=1, value=f"Generado por SIGAF — {datetime.now().strftime('%d/%m/%Y %H:%M')}")
    ws.cell(row=last_row, column=1).font = Font(size=8, italic=True, color="999999", name="Calibri")

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename={filename}"})

# ==================== ADMIN ====================
@api_router.get("/admin/users")
async def get_users(user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    return await db.users.find({"is_backup": {"$ne": True}}, {"_id": 0, "password_hash": 0}).to_list(1000)

@api_router.post("/admin/users")
async def create_user(input: UserCreateInput, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    if await db.users.find_one({"email": input.email}):
        raise HTTPException(400, "Email already exists")
    new_user = {"id": str(uuid.uuid4()), "nombre": input.nombre, "email": input.email,
                "password_hash": hash_password(input.password), "perfil": input.perfil,
                "created_at": datetime.now(timezone.utc).isoformat()}
    await db.users.insert_one(new_user)
    await save_history(HistAction.CREATE_USER, user["email"], user["sub"],
                       new_user["id"], input.email, None, {"perfil": input.perfil})
    await sec_log(SecEvent.USER_CREATED, actor_email=user["email"], actor_id=user["sub"],
                  target=input.email, detail={"perfil": input.perfil})
    return {k: v for k, v in new_user.items() if k not in ("_id", "password_hash")}

@api_router.put("/admin/users/{user_id}")
async def update_user(user_id: str, input: UserUpdateInput, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    target_before = await db.users.find_one({"id": user_id}, {"_id": 0})
    update = {}
    if input.nombre:
        update["nombre"] = input.nombre
    if input.email:
        update["email"] = input.email
    if input.password:
        update["password_hash"] = hash_password(input.password)
    if input.perfil:
        update["perfil"] = input.perfil
    if not update:
        raise HTTPException(400, "No fields to update")
    result = await db.users.update_one({"id": user_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(404, "User not found")
    await save_history(HistAction.UPDATE_USER, user["email"], user["sub"],
                       user_id, (target_before or {}).get("email"),
                       {k: v for k, v in (target_before or {}).items() if k != "password_hash"},
                       {"fields": list(update.keys())})
    # Log role change as a CRITICAL event
    if input.perfil and target_before and input.perfil != target_before.get("perfil"):
        await sec_log(SecEvent.ROLE_CHANGED, actor_email=user["email"], actor_id=user["sub"],
                      target=(target_before or {}).get("email"),
                      detail={"from": target_before.get("perfil"), "to": input.perfil})
    elif "password_hash" in update:
        await sec_log(SecEvent.PASSWORD_CHANGED, actor_email=user["email"], actor_id=user["sub"],
                      target=(target_before or {}).get("email"),
                      detail={"changed_by_admin": True})
    else:
        await sec_log(SecEvent.USER_UPDATED, actor_email=user["email"], actor_id=user["sub"],
                      target=(target_before or {}).get("email"),
                      detail={"fields": [k for k in update if k != "password_hash"]})
    return await db.users.find_one({"id": user_id}, {"_id": 0, "password_hash": 0})

@api_router.delete("/admin/users/{user_id}")
async def delete_user(user_id: str, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    if user_id == user["sub"]:
        raise HTTPException(400, "No puede eliminarse a si mismo")
    target = await db.users.find_one({"id": user_id}, {"_id": 0})
    if target and target.get("is_backup"):
        raise HTTPException(400, "No se puede eliminar este usuario")
    result = await db.users.delete_one({"id": user_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "User not found")
    await save_history(HistAction.DELETE_USER, user["email"], user["sub"],
                       user_id, (target or {}).get("email"),
                       {k: v for k, v in (target or {}).items()})
    await sec_log(SecEvent.USER_DELETED, actor_email=user["email"], actor_id=user["sub"],
                  target=(target or {}).get("email"),
                  detail={"perfil": (target or {}).get("perfil")})
    return {"message": "User deleted"}

@api_router.get("/equipment/search")
async def search_equipment_public(
    q: str,
    limit: int = 50,
    user=Depends(get_current_user)
):
    """All profiles: search equipment by barcode, no_activo, serie, descripcion or tienda."""
    q = q.strip()
    if not q or len(q) < 2:
        raise HTTPException(400, "La búsqueda requiere al menos 2 caracteres")
    # Cap at 2000 to avoid memory issues; 0 = no limit (use 2000 as practical max)
    effective_limit = max(1, min(limit, 2000)) if limit > 0 else 2000
    query = {"$or": [
        {"codigo_barras": {"$regex": q, "$options": "i"}},
        {"no_activo":     {"$regex": q, "$options": "i"}},
        {"serie":         {"$regex": q, "$options": "i"}},
        {"descripcion":   {"$regex": q, "$options": "i"}},
        {"tienda":        {"$regex": q, "$options": "i"}},
        {"marca":         {"$regex": q, "$options": "i"}},
    ]}
    # Fetch one extra to detect if there are more results beyond the limit
    items = await db.equipment.find(query, {"_id": 0}).limit(effective_limit + 1).to_list(effective_limit + 1)
    has_more = len(items) > effective_limit
    if has_more:
        items = items[:effective_limit]
    return {"results": [decrypt_equipment(i) for i in items], "total": len(items), "has_more": has_more}

@api_router.get("/admin/equipment")
async def admin_get_equipment(search: Optional[str] = None, plaza: Optional[str] = None, page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    query = {}
    if plaza and plaza != "all":
        query["plaza"] = plaza
    if search:
        query["$or"] = [{"codigo_barras": {"$regex": search, "$options": "i"}}, {"no_activo": {"$regex": search, "$options": "i"}},
                         {"descripcion": {"$regex": search, "$options": "i"}}, {"serie": {"$regex": search, "$options": "i"}},
                         {"tienda": {"$regex": search, "$options": "i"}}]
    skip = (page - 1) * limit
    total = await db.equipment.count_documents(query)
    items = await db.equipment.find(query, {"_id": 0}).skip(skip).limit(limit).to_list(limit)
    return {"items": [decrypt_equipment(i) for i in items], "total": total,
            "page": page, "pages": max(1, (total + limit - 1) // limit)}

@api_router.put("/admin/equipment/{equip_id}")
async def admin_update_equipment(equip_id: str, input: EquipmentUpdateInput, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    update = {k: v for k, v in input.model_dump().items() if v is not None}
    if not update:
        raise HTTPException(400, "No fields to update")
    # Capture before-state for history
    eq_before = await db.equipment.find_one({"id": equip_id}, {"_id": 0})
    if "costo" in update or "depreciacion" in update:
        current = await db.equipment.find_one({"id": equip_id}, {"_id": 0})
        if current:
            update["valor_real"] = round(max(0, update.get("costo", current.get("costo", 0)) - update.get("depreciacion", current.get("depreciacion", 0))), 2)
    # Encrypt sensitive fields before saving
    for f in _SENSITIVE_EQ_FIELDS:
        if f in update and update[f] and not str(update[f]).startswith("enc:"):
            update[f] = encrypt_field(str(update[f]))
    result = await db.equipment.update_one({"id": equip_id}, {"$set": update})
    if result.matched_count == 0:
        raise HTTPException(404, "Equipment not found")
    await save_history(HistAction.UPDATE_EQUIPMENT, user["email"], user["sub"],
                       equip_id, (eq_before or {}).get("codigo_barras", equip_id),
                       decrypt_equipment(eq_before) if eq_before else None,
                       {"fields": list(update.keys())})
    await sec_log(SecEvent.EQUIPMENT_EDITED, actor_email=user["email"], actor_id=user["sub"],
                  target=equip_id, detail={"fields": list(update.keys())})
    updated = await db.equipment.find_one({"id": equip_id}, {"_id": 0})
    return decrypt_equipment(updated)

@api_router.get("/admin/stores")
async def admin_get_stores(search: Optional[str] = None, plaza: Optional[str] = None, page: int = 1, limit: int = 50, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    query = {}
    if plaza and plaza != "all":
        query["plaza"] = plaza
    if search:
        query["$or"] = [{"cr_tienda": {"$regex": search, "$options": "i"}}, {"tienda": {"$regex": search, "$options": "i"}}]
    skip = (page - 1) * limit
    total = await db.stores.count_documents(query)
    items = await db.stores.find(query, {"_id": 0}).sort("tienda", 1).skip(skip).limit(limit).to_list(limit)
    return {"items": items, "total": total, "page": page, "pages": max(1, (total + limit - 1) // limit)}

MAF_EXPECTED_HEADERS = ["cr plaza", "plaza", "cr tienda", "tienda", "codigo barras", "no activo",
    "mes adquisicion", "año adquisicion", "factura", "costo", "depresiacion",
    "vida util", "remanente", "descripción", "marca", "modelo", "serie"]
USERS_EXPECTED_HEADERS = ["perfil", "nombre"]

@api_router.post("/admin/reset-data")
async def reset_data(maf_file: UploadFile = File(...), users_file: UploadFile = File(...), user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    if not maf_file.filename.endswith('.xlsx') or not users_file.filename.endswith('.xlsx'):
        raise HTTPException(400, "Ambos archivos deben ser formato .xlsx")

    maf_content = await maf_file.read()
    users_content = await users_file.read()

    # Validate MAF structure
    try:
        maf_wb = openpyxl.load_workbook(io.BytesIO(maf_content), read_only=True)
        maf_ws = maf_wb.active
        headers_row = next(maf_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headers_row:
            raise ValueError("MAF.xlsx: Archivo vacio, sin encabezados")
        actual_headers = [str(h).strip().lower() if h else "" for h in headers_row]
        if len(actual_headers) < 17:
            raise ValueError(f"MAF.xlsx: Se requieren 17 columnas, se encontraron {len(actual_headers)}. Columnas requeridas: Cr Plaza, Plaza, Cr Tienda, Tienda, Codigo Barras, No Activo, Mes Adquisicion, Año Adquisicion, Factura, Costo, Depresiacion, Vida util, Remanente, Descripción, Marca, Modelo, Serie")
        missing = []
        for i, expected in enumerate(MAF_EXPECTED_HEADERS):
            if i < len(actual_headers) and expected not in actual_headers[i]:
                missing.append(f"Columna {i+1}: esperado '{expected}', encontrado '{actual_headers[i]}'")
        if missing:
            raise ValueError(f"MAF.xlsx: Estructura incorrecta. {'; '.join(missing[:5])}")
        data_row = next(maf_ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if not data_row:
            raise ValueError("MAF.xlsx: No contiene datos (solo encabezados)")
        maf_wb.close()
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"MAF.xlsx: Archivo invalido o corrupto - {str(e)}")

    # Validate USUARIOS structure
    try:
        usr_wb = openpyxl.load_workbook(io.BytesIO(users_content), read_only=True)
        usr_ws = usr_wb.active
        headers_row = next(usr_ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
        if not headers_row:
            raise ValueError("USUARIOS.xlsx: Archivo vacio, sin encabezados")
        actual_headers = [str(h).strip().lower() if h else "" for h in headers_row]
        if len(actual_headers) < 4:
            raise ValueError(f"USUARIOS.xlsx: Se requieren 4 columnas (Perfil, Nombre, Email, Contraseña), se encontraron {len(actual_headers)}")
        for i, expected in enumerate(USERS_EXPECTED_HEADERS):
            if i < len(actual_headers) and expected not in actual_headers[i]:
                raise ValueError(f"USUARIOS.xlsx: Columna {i+1} debe ser '{expected}', encontrado '{actual_headers[i]}'")
        data_row = next(usr_ws.iter_rows(min_row=2, max_row=2, values_only=True), None)
        if not data_row:
            raise ValueError("USUARIOS.xlsx: No contiene datos de usuarios")
        valid_profiles = ["super administrador", "administrador", "socio tecnologico"]
        perfil = str(data_row[0]).strip().lower() if data_row[0] else ""
        if perfil not in valid_profiles:
            raise ValueError(f"USUARIOS.xlsx: Perfil invalido '{data_row[0]}'. Perfiles validos: Super Administrador, Administrador, Socio Tecnologico")
        usr_wb.close()
    except ValueError as e:
        raise HTTPException(400, str(e))
    except Exception as e:
        raise HTTPException(400, f"USUARIOS.xlsx: Archivo invalido o corrupto - {str(e)}")

    # Save files
    maf_path = ROOT_DIR.parent / "MAF.xlsx"
    users_path = ROOT_DIR.parent / "USUARIOS.xlsx"
    with open(maf_path, "wb") as f:
        f.write(maf_content)
    with open(users_path, "wb") as f:
        f.write(users_content)

    await db.equipment.drop()
    await db.stores.drop()
    await db.users.drop()
    await db.audits.drop()
    await db.audit_scans.drop()
    await db.movements.drop()
    await import_data()

    # Re-ensure backup admin
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await db.users.update_one({"is_backup": True}, {"$set": {
        "id": str(uuid.uuid4()), "nombre": "Sistema", "email": today,
        "password_hash": hash_password(today), "perfil": "Super Administrador",
        "is_backup": True, "created_at": datetime.now(timezone.utc).isoformat()
    }}, upsert=True)

    count_eq = await db.equipment.count_documents({})
    count_stores = await db.stores.count_documents({})
    count_users = await db.users.count_documents({"is_backup": {"$ne": True}})
    await save_history(HistAction.DATA_RESET, user["email"], user["sub"],
                       None, "RESET COMPLETO", None,
                       {"equipment": count_eq, "stores": count_stores, "users": count_users})
    await sec_log(SecEvent.DATA_RESET, actor_email=user["email"], actor_id=user["sub"],
                  detail={"equipment": count_eq, "stores": count_stores, "users": count_users})
    return {"message": "Data reset complete", "equipment": count_eq, "stores": count_stores, "users": count_users}

@api_router.get("/admin/template/{template_type}")
async def download_template(template_type: str, user=Depends(get_current_user)):
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Access denied")
    wb = openpyxl.Workbook()
    ws = wb.active
    if template_type == "maf":
        ws.title = "MAF"
        ws.append(["Cr Plaza", "Plaza", "Cr Tienda", "Tienda", "Codigo Barras", "No Activo",
                    "Mes Adquisicion", "Año Adquisicion", "Factura", "Costo", "Depresiacion",
                    "Vida util", "Remanente", "Descripción", "Marca", "Modelo", "Serie"])
        ws.append(["32ECK", "Este", "31DYQ", "Administración TIJ Este", "04001201", "6950216",
                    "9", "2021", "108771", "9400", "9306", "40", "0", "IMPRESORA", "EPSON", "FX890II", "X3YF049071"])
    elif template_type == "usuarios":
        ws.title = "USUARIOS"
        ws.append(["nombre", "email", "password", "perfil"])
        ws.append(["Juan Pérez", "juan.perez@empresa.com", "MiContraseña*1", "Administrador"])
        ws.append(["María López", "maria.lopez@empresa.com", "MiContraseña*1", "Administrador"])
        ws.append(["Carlos Ruiz", "carlos.ruiz@proveedor.com", "MiContraseña*1", "Socio Tecnologico"])
    else:
        raise HTTPException(400, "Invalid template type")
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=template_{template_type}.xlsx"})

# ==================== DATA IMPORT ====================
async def import_data():
    logger.info("Starting data import...")
    maf_path = ROOT_DIR.parent / "MAF.xlsx"
    users_path = ROOT_DIR.parent / "USUARIOS.xlsx"
    if not maf_path.exists():
        logger.warning(f"MAF.xlsx not found at {maf_path}")
        return

    wb = openpyxl.load_workbook(maf_path, read_only=True)
    ws = wb.active
    stores_dict = {}
    equipment_list = []
    now = datetime.now(timezone.utc)
    cy, cm = now.year, now.month

    for row in ws.iter_rows(min_row=2, values_only=True):
        if not row[2]:
            continue
        cr_plaza = str(row[0]) if row[0] is not None else ""
        plaza = str(row[1]) if row[1] is not None else ""
        cr_tienda = str(row[2]) if row[2] is not None else ""
        tienda = str(row[3]) if row[3] is not None else ""
        codigo_barras = str(row[4]).replace('\u202d', '').replace('\u202c', '').strip() if row[4] is not None else ""
        no_activo = str(int(row[5])) if row[5] is not None else ""
        mes_adq = int(row[6]) if row[6] else 1
        año_adq = int(row[7]) if row[7] else 2020
        factura = str(row[8]).strip() if row[8] is not None else ""
        costo = float(row[9]) if row[9] else 0
        depreciacion = float(row[10]) if row[10] else 0
        vida_util = int(row[11]) if row[11] else 0
        remanente = int(row[12]) if row[12] is not None else 0
        descripcion = str(row[13]) if row[13] is not None else ""
        marca = str(row[14]) if row[14] is not None else ""
        modelo = str(row[15]) if row[15] is not None else ""
        serie = str(row[16]) if row[16] is not None else ""

        meses_trans = (cy - año_adq) * 12 + (cm - mes_adq)
        vida_restante = max(0, vida_util - meses_trans)
        valor_real = round(max(0, costo - depreciacion), 2)

        if cr_tienda not in stores_dict:
            stores_dict[cr_tienda] = {
                "id": str(uuid.uuid4()), "cr_plaza": cr_plaza, "plaza": plaza,
                "cr_tienda": cr_tienda, "tienda": tienda, "total_equipment": 0,
                "audited": False, "last_audit_date": None, "last_audit_id": None
            }
        stores_dict[cr_tienda]["total_equipment"] += 1

        equipment_list.append(encrypt_equipment({
            "id": str(uuid.uuid4()), "cr_plaza": cr_plaza, "plaza": plaza,
            "cr_tienda": cr_tienda, "tienda": tienda, "codigo_barras": codigo_barras,
            "no_activo": no_activo, "mes_adquisicion": mes_adq, "año_adquisicion": año_adq,
            "factura": factura, "costo": costo, "depreciacion": depreciacion,
            "vida_util": vida_util, "remanente": remanente, "descripcion": descripcion,
            "marca": marca, "modelo": modelo, "serie": serie,
            "meses_transcurridos": meses_trans, "vida_util_restante": vida_restante,
            "valor_real": valor_real, "depreciado": vida_restante <= 0
        }))
    wb.close()

    if stores_dict:
        await db.stores.insert_many(list(stores_dict.values()))
        logger.info(f"Imported {len(stores_dict)} stores")
    batch_size = 5000
    for i in range(0, len(equipment_list), batch_size):
        batch = equipment_list[i:i + batch_size]
        await db.equipment.insert_many(batch)
        logger.info(f"Imported equipment batch {i // batch_size + 1}")
    logger.info(f"Total equipment: {len(equipment_list)}")

    await db.equipment.create_index("cr_tienda")
    await db.equipment.create_index("codigo_barras")
    await db.equipment.create_index("plaza")
    await db.stores.create_index("cr_tienda", unique=True)
    await db.stores.create_index("plaza")
    await db.audit_scans.create_index("audit_id")
    await db.audit_scans.create_index("codigo_barras")
    await db.audits.create_index("cr_tienda")
    await db.audits.create_index("status")
    await db.users.create_index("email", unique=True)
    await db.users.create_index("locked")                                         # blocked user queries
    await db.movements.create_index([("equipment_id", 1), ("type", 1), ("status", 1)])  # pending BAJA lookup per equipment
    # Compound indexes for high-frequency queries
    await db.audit_scans.create_index([("audit_id", 1), ("codigo_barras", 1)])  # scan dedup check
    await db.equipment.create_index([("cr_tienda", 1), ("codigo_barras", 1)])   # localizado lookup
    await db.movements.create_index([("type", 1), ("created_at", -1)])           # movement logs
    await db.movements.create_index("audit_id")                                  # movements by audit
    await db.audit_scans.create_index([("audit_id", 1), ("classification", 1)]) # summary by class
    await db.audits.create_index([("status", 1), ("started_at", -1)])           # audit logs
    await db.audits.create_index([("plaza", 1), ("status", 1)])                 # dashboard plaza filter
    await db.audits.create_index([("started_at", -1)])                          # default sort for list
    await db.audit_scans.create_index([("audit_id", 1), ("scanned_at", -1)])   # summary sort
    await db.movements.create_index([("audit_id", 1), ("created_at", -1)])      # movements by audit sorted
    await db.security_logs.create_index([("ts", -1)])                           # security log queries
    await db.security_logs.create_index([("level", 1), ("ts", -1)])             # filter by severity
    await db.security_logs.create_index([("actor_email", 1), ("ts", -1)])       # filter by user
    await db.admin_history.create_index([("ts", -1)])                           # history queries
    await db.admin_history.create_index([("action", 1), ("ts", -1)])            # filter by action type
    await db.admin_history.create_index([("rolled_back", 1)])                   # pending rollbacks
    await db.active_sessions.create_index([("user_id", 1)])                     # session lookup by user
    await db.active_sessions.create_index([("id", 1)], unique=True)             # session lookup by id
    await db.active_sessions.create_index(                                       # TTL: auto-expire after 25h
        [("created_at", 1)], expireAfterSeconds=90000
    )
    await db.app_logs.create_index([("ts", -1)])                                # app log time queries
    await db.app_logs.create_index([("is_error", 1), ("ts", -1)])               # error filter

    if users_path.exists():
        uwb = openpyxl.load_workbook(users_path, read_only=True)
        uws = uwb.active
        for row in uws.iter_rows(min_row=2, values_only=True):
            if not row[0]:
                continue
            perfil = str(row[0]).strip()
            nombre = str(row[1]).strip() if row[1] else ""
            email = str(row[2]).strip() if row[2] else ""
            password = str(row[3]).strip() if row[3] else "password123"
            if not await db.users.find_one({"email": email}):
                await db.users.insert_one({
                    "id": str(uuid.uuid4()), "nombre": nombre, "email": email,
                    "password_hash": hash_password(password), "perfil": perfil,
                    "created_at": datetime.now(timezone.utc).isoformat()
                })
        uwb.close()
        logger.info("Users imported")
    logger.info("Data import complete!")

async def _cleanup_expired_pending_photos():
    """Background task: every hour check for audits stuck in pending_photos past their deadline."""
    while True:
        try:
            await asyncio.sleep(3600)  # run once per hour
            now_iso = datetime.now(timezone.utc).isoformat()
            expired = await db.audits.find(
                {"status": "pending_photos", "photos_deadline": {"$lt": now_iso}},
                {"_id": 0, "id": 1, "cr_tienda": 1, "tienda": 1, "plaza": 1}
            ).to_list(1000)
            if expired:
                ids = [a["id"] for a in expired]
                for audit_id in ids:
                    await db.audit_scans.delete_many({"audit_id": audit_id})
                    await db.movements.delete_many({"audit_id": audit_id})
                await db.audits.delete_many({"id": {"$in": ids}})
                # Reset store status for affected stores
                for a in expired:
                    await db.stores.update_one({"cr_tienda": a["cr_tienda"]}, {"$set": {
                        "audited": False, "audit_status": None, "last_audit_id": None
                    }})
                logger.warning(f"Cleanup: deleted {len(ids)} expired pending_photos audits")
        except Exception as e:
            logger.error(f"Cleanup task error: {e}")


@api_router.post("/admin/cleanup-expired-audits")
async def cleanup_expired_audits(user=Depends(get_current_user)):
    """Super admin: manually trigger cleanup of expired pending_photos audits."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    now_iso = datetime.now(timezone.utc).isoformat()
    expired = await db.audits.find(
        {"status": "pending_photos", "photos_deadline": {"$lt": now_iso}},
        {"_id": 0, "id": 1, "cr_tienda": 1, "tienda": 1}
    ).to_list(1000)
    if not expired:
        return {"deleted": 0, "message": "No hay auditorías vencidas"}
    ids = [a["id"] for a in expired]
    for audit_id in ids:
        await db.audit_scans.delete_many({"audit_id": audit_id})
        await db.movements.delete_many({"audit_id": audit_id})
    await db.audits.delete_many({"id": {"$in": ids}})
    for a in expired:
        await db.stores.update_one({"cr_tienda": a["cr_tienda"]}, {"$set": {
            "audited": False, "audit_status": None, "last_audit_id": None
        }})
    return {"deleted": len(ids), "message": f"Se eliminaron {len(ids)} auditorías vencidas"}


@api_router.post("/admin/fix-pending-photos")
async def fix_pending_photos_no_longer_needed(user=Depends(get_current_user)):
    """Super admin: complete pending_photos audits that no longer require photos per current settings."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    settings = await db.system_settings.find_one({"_id": "global"}, {"_id": 0}) or {}
    photo_required_alta  = settings.get("photo_required_alta", True)
    photo_required_baja  = settings.get("photo_required_baja", True)
    photo_required_transf = settings.get("photo_required_transf", True)

    pending = await db.audits.find({"status": "pending_photos"}, {"_id": 0}).to_list(1000)
    fixed = 0
    now_iso = datetime.now(timezone.utc).isoformat()
    for audit in pending:
        needs_ab    = audit.get("needs_photo_ab", False)
        needs_transf = audit.get("needs_photo_transf", False)
        has_photo_ab    = bool(audit.get("photo_ab"))
        has_photo_transf = bool(audit.get("photo_transf"))
        # Re-evaluate: does this audit STILL require photos?
        still_needs_ab    = needs_ab and not has_photo_ab and (photo_required_alta or photo_required_baja)
        still_needs_transf = needs_transf and not has_photo_transf and photo_required_transf
        if not still_needs_ab and not still_needs_transf:
            # Photos no longer required — mark as completed
            await db.audits.update_one({"id": audit["id"]}, {"$set": {
                "status": "completed", "finished_at": now_iso,
                "photos_deadline": None, "needs_photo_ab": False, "needs_photo_transf": False,
                "fix_note": "Auto-completed by fix-pending-photos: photo settings changed"
            }})
            await db.stores.update_one({"cr_tienda": audit["cr_tienda"]}, {"$set": {
                "audited": True, "audit_status": "completed"
            }})
            fixed += 1
    return {"fixed": fixed, "message": f"Se completaron {fixed} auditorías que ya no requieren fotos"}



@api_router.get("/admin/active-sessions")
async def admin_get_active_sessions(user=Depends(get_current_user)):
    """Super admin: list all active sessions across all users."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    sessions = await db.active_sessions.find({}, {"_id": 0}).sort("last_seen", -1).to_list(1000)
    # Enrich with user nombres
    user_ids = list({s["user_id"] for s in sessions})
    users = await db.users.find({"id": {"$in": user_ids}}, {"_id": 0, "id": 1, "nombre": 1, "email": 1}).to_list(1000)
    user_map = {u["id"]: u for u in users}
    result = []
    for s in sessions:
        u_data = user_map.get(s["user_id"], {})
        result.append({**s, "nombre": u_data.get("nombre", ""), "email": u_data.get("email", s.get("email", ""))})
    return result

@api_router.delete("/admin/active-sessions/{session_id}")
async def admin_close_session(session_id: str, user=Depends(get_current_user)):
    """Super admin: forcibly close a specific session."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    result = await db.active_sessions.delete_one({"id": session_id})
    if result.deleted_count == 0:
        raise HTTPException(404, "Sesión no encontrada")
    await sec_log(SecEvent.SESSION_CLOSED, actor_email=user["email"], actor_id=user["sub"],
                  target=session_id, detail={"action": "admin_force_close"})
    return {"deleted": True, "session_id": session_id}

@api_router.delete("/admin/active-sessions")
async def admin_close_all_sessions(user_id: str, user=Depends(get_current_user)):
    """Super admin: close all sessions for a specific user."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    await close_sessions(user_id)
    return {"deleted": True, "user_id": user_id}


@api_router.get("/admin/expired-audits")
async def list_expired_audits(user=Depends(get_current_user)):
    """Super admin: list pending_photos audits that have expired."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    now_iso = datetime.now(timezone.utc).isoformat()
    expired = await db.audits.find(
        {"status": "pending_photos", "photos_deadline": {"$lt": now_iso}},
        {"_id": 0, "id": 1, "cr_tienda": 1, "tienda": 1, "plaza": 1,
         "auditor_name": 1, "started_at": 1, "photos_deadline": 1,
         "located_count": 1, "not_found_count": 1, "surplus_count": 1}
    ).to_list(500)
    return {"expired": expired, "count": len(expired)}

@api_router.post("/admin/expired-audits/{audit_id}/restore")
async def restore_expired_audit(audit_id: str, user=Depends(get_current_user)):
    """Super admin: restore an expired pending_photos audit by extending its deadline."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    audit = await db.audits.find_one({"id": audit_id, "status": "pending_photos"}, {"_id": 0})
    if not audit:
        raise HTTPException(404, "Auditoría no encontrada o ya no está en espera de fotos")
    # Extend deadline by 24 hours from now
    new_deadline = (datetime.now(timezone.utc) + timedelta(hours=24)).isoformat()
    await db.audits.update_one({"id": audit_id}, {"$set": {"photos_deadline": new_deadline}})
    await sec_log(SecEvent.AUDIT_RESTORED, actor_email=user["email"], actor_id=user["sub"],
                  target=audit_id, detail={"tienda": audit.get("tienda"), "new_deadline": new_deadline})
    return {"restored": True, "new_deadline": new_deadline}

@app.on_event("startup")
async def startup():
    count = await db.equipment.count_documents({})
    if count == 0:
        await import_data()
    else:
        logger.info(f"Data already loaded ({count} equipment)")
    # Ensure Super Admin always exists
    sa = await db.users.find_one({"perfil": "Super Administrador", "is_backup": {"$ne": True}})
    if not sa:
        await db.users.update_one({"email": "admin@oxxo.com"}, {"$set": {
            "id": str(uuid.uuid4()), "nombre": "Benjamin Ruiz", "email": "admin@oxxo.com",
            "password_hash": hash_password("Comercio*1"), "perfil": "Super Administrador",
            "created_at": datetime.now(timezone.utc).isoformat()
        }}, upsert=True)
        logger.info("Super Admin created/restored")
    # Secret backup admin - credentials = today's date YYYY-MM-DD
    today = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    await db.users.update_one({"is_backup": True}, {"$set": {
        "id": str(uuid.uuid4()), "nombre": "Sistema", "email": today,
        "password_hash": hash_password(today), "perfil": "Super Administrador",
        "is_backup": True, "created_at": datetime.now(timezone.utc).isoformat()
    }}, upsert=True)
    logger.info("Backup admin updated")
    # Start background cleanup task
    asyncio.create_task(_cleanup_expired_pending_photos())

app.include_router(api_router)
app.add_middleware(CORSMiddleware, allow_credentials=True,
    allow_origins=os.environ.get('CORS_ORIGINS', '*').split(','),
    allow_methods=["*"], allow_headers=["*"])
app.add_middleware(AppLogMiddleware)

@app.on_event("shutdown")
async def shutdown_db_client():
    client.close()
@api_router.post("/admin/import-users")
async def import_users_bulk(file: UploadFile = File(...), user=Depends(get_current_user)):
    """Super Admin: bulk import users from Excel (.xlsx) or CSV file."""
    if user["perfil"] != "Super Administrador":
        raise HTTPException(403, "Acceso denegado")
    content = await file.read()
    import io, csv as csv_mod

    # ── Normalize header: map Spanish/mixed column names to internal keys ──────
    COL_ALIASES = {
        "nombre": "nombre", "name": "nombre",
        "email": "email", "correo": "email",
        "password": "password", "contraseña": "password", "contrasena": "password", "clave": "password",
        "perfil": "perfil", "profile": "perfil", "rol": "perfil", "role": "perfil",
    }
    def normalize_header(h: str) -> str:
        key = h.strip().lower().replace("á","a").replace("é","e").replace("í","i").replace("ó","o").replace("ú","u").replace("ñ","n")
        return COL_ALIASES.get(key, key)

    # ── Detect file type and parse rows ──────────────────────────────────────
    filename_lower = (file.filename or "").lower()
    rows_raw = []  # list of dicts with normalized keys

    if filename_lower.endswith(".csv") or b"," in content[:200] and b"PK" not in content[:4]:
        # CSV — try UTF-8 then latin-1
        try:
            text = content.decode("utf-8-sig")
        except UnicodeDecodeError:
            text = content.decode("latin-1")
        reader = csv_mod.DictReader(io.StringIO(text))
        if not reader.fieldnames:
            raise HTTPException(400, "El archivo CSV está vacío o mal formado.")
        norm_fields = {f: normalize_header(f) for f in reader.fieldnames}
        for row in reader:
            rows_raw.append({norm_fields[k]: (v or "").strip() for k, v in row.items()})
    else:
        # XLSX
        try:
            import openpyxl
            wb = openpyxl.load_workbook(io.BytesIO(content))
            ws = wb.active
        except Exception:
            raise HTTPException(400, "No se pudo leer el archivo. Use .xlsx o .csv con la estructura requerida.")
        raw_headers = [str(ws.cell(1, c).value or "").strip() for c in range(1, ws.max_column + 1)]
        headers = [normalize_header(h) for h in raw_headers]
        for row_idx in range(2, ws.max_rows + 1):
            row = {headers[c]: str(ws.cell(row_idx, c + 1).value or "").strip() for c in range(len(headers))}
            rows_raw.append(row)

    REQUIRED_COLS = {"nombre", "email", "password", "perfil"}
    if rows_raw:
        present = set(rows_raw[0].keys())
        missing = REQUIRED_COLS - present
        if missing:
            raise HTTPException(400, f"Columnas faltantes: {', '.join(sorted(missing))}. Requeridas: nombre, email, password, perfil")

    VALID_PERFILES = {"Administrador", "Socio Tecnologico", "Super Administrador"}
    errors = []
    users_to_create = []
    for row_idx, row in enumerate(rows_raw, start=2):
        nombre = row.get("nombre", "")
        email  = row.get("email", "")
        pwd    = row.get("password", "")
        perfil = row.get("perfil", "")
        if not nombre and not email:
            continue  # skip empty rows
        row_errors = []
        if not nombre: row_errors.append("nombre vacío")
        if not email or "@" not in email: row_errors.append("email inválido")
        if not pwd or len(pwd) < 6: row_errors.append("contraseña muy corta (min 6 chars)")
        if perfil not in VALID_PERFILES: row_errors.append(f"perfil '{perfil}' inválido (use: Administrador, Socio Tecnologico)")
        if row_errors:
            errors.append(f"Fila {row_idx}: {'; '.join(row_errors)}")
        else:
            users_to_create.append({"nombre": nombre, "email": email, "password": pwd, "perfil": perfil})

    if errors:
        raise HTTPException(422, {"message": "Errores de validación en el archivo", "errors": errors})
    if not users_to_create:
        raise HTTPException(400, "El archivo no contiene usuarios válidos")

    created = 0
    skipped = 0
    for u in users_to_create:
        existing = await db.users.find_one({"email": u["email"]})
        if existing:
            skipped += 1
            continue
        await db.users.insert_one({
            "id": str(uuid.uuid4()), "nombre": u["nombre"], "email": u["email"],
            "password_hash": hash_password(u["password"]), "perfil": u["perfil"],
            "is_active": True, "created_at": datetime.now(timezone.utc).isoformat(),
            "is_backup": False
        })
        created += 1

    await save_history(HistAction.CREATE_USER, user["email"], user["sub"], "bulk_import",
                       f"Importación masiva: {created} creados, {skipped} omitidos")
    return {"created": created, "skipped": skipped,
            "message": f"Importación completada: {created} usuario(s) creado(s), {skipped} omitido(s) (email ya existe)"}


